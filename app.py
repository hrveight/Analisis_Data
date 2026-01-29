import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from scipy import stats
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ==================== CONFIG ====================
st.set_page_config(
    page_title="Dashboard Analisis PSDKP",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== STYLING ====================
st.markdown("""
<style>
    .main {padding: 1rem 2rem;}
    .stMetric {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
    }
    .stMetric label {color: white !important;}
    .stMetric [data-testid="stMetricValue"] {color: white !important;}
    .info-box {
        background-color: #e7f3ff;
        border-left: 4px solid #2196f3;
        padding: 1rem;
        border-radius: 4px;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border-left: 4px solid #ffc107;
        padding: 1rem;
        border-radius: 4px;
        margin: 1rem 0;
    }
    .success-box {
        background-color: #d4edda;
        border-left: 4px solid #28a745;
        padding: 1rem;
        border-radius: 4px;
        margin: 1rem 0;
    }
    .assumption-box {
        background-color: #f8f9fa;
        border: 1px solid #dee2e6;
        padding: 1rem;
        border-radius: 4px;
        margin: 0.5rem 0;
    }
    .download-section {
        background-color: #f0f2f6;
        border: 2px dashed #667eea;
        padding: 1.5rem;
        border-radius: 8px;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# ==================== SESSION STATE ====================
if 'df' not in st.session_state:
    st.session_state.df = None
if 'df_raw' not in st.session_state:
    st.session_state.df_raw = None
if 'column_types' not in st.session_state:
    st.session_state.column_types = {}
if 'analysis_results' not in st.session_state:
    st.session_state.analysis_results = None
if 'selected_analysis' not in st.session_state:
    st.session_state.selected_analysis = None
if 'all_analysis_history' not in st.session_state:
    st.session_state.all_analysis_history = []

# ==================== HELPER FUNCTIONS ====================
def read_file(uploaded_file):
    """Baca file CSV atau Excel dengan handling error"""
    try:
        name = uploaded_file.name.lower()
        if name.endswith('.csv'):
            raw = uploaded_file.getvalue()
            try:
                df = pd.read_csv(io.BytesIO(raw))
            except UnicodeDecodeError:
                df = pd.read_csv(io.BytesIO(raw), encoding='latin-1')
            except:
                df = pd.read_csv(io.BytesIO(raw), sep=';')
            return df, None
        elif name.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(uploaded_file, engine='openpyxl')
            return df, None
        else:
            return None, "Format tidak didukung"
    except Exception as e:
        return None, f"Error: {str(e)}"

def infer_column_type(series):
    """Deteksi otomatis tipe kolom"""
    if pd.api.types.is_bool_dtype(series):
        return 'Boolean'
    
    if pd.api.types.is_datetime64_any_dtype(series):
        return 'Tanggal'
    
    if pd.api.types.is_numeric_dtype(series):
        return 'Angka'
    
    if series.dtype == object:
        sample = series.dropna().head(20)
        if len(sample) > 0:
            try:
                parsed = pd.to_datetime(sample, errors='coerce', dayfirst=True)
                if parsed.notna().mean() >= 0.7:
                    return 'Tanggal'
            except:
                pass
    
    return 'Teks/Kategori'

def apply_column_types(df, type_mapping):
    """Konversi tipe kolom sesuai pilihan user"""
    warnings = []
    df_new = df.copy()
    
    for col, dtype in type_mapping.items():
        try:
            if dtype == 'Angka':
                df_new[col] = pd.to_numeric(df_new[col], errors='coerce')
                null_pct = df_new[col].isna().mean() * 100
                if null_pct > 30:
                    warnings.append(f"Kolom '{col}': {null_pct:.0f}% gagal dikonversi ke angka")
            
            elif dtype == 'Tanggal':
                df_new[col] = pd.to_datetime(df_new[col], errors='coerce', dayfirst=True)
                null_pct = df_new[col].isna().mean() * 100
                if null_pct > 30:
                    warnings.append(f"Kolom '{col}': {null_pct:.0f}% gagal dikonversi ke tanggal")
            
            elif dtype == 'Teks/Kategori':
                df_new[col] = df_new[col].astype(str)
            
            elif dtype == 'Boolean':
                s = df_new[col]
                if pd.api.types.is_numeric_dtype(s):
                    df_new[col] = s.astype('Int64').map({1: True, 0: False})
                else:
                    ss = s.astype(str).str.lower().str.strip()
                    df_new[col] = ss.map({
                        'true': True, 'false': False, '1': True, '0': False,
                        'ya': True, 'tidak': False, 'y': True, 'n': False
                    })
        except Exception as e:
            warnings.append(f"Gagal konversi '{col}': {str(e)}")
    
    return df_new, warnings

def get_columns_by_type(df, type_mapping):
    """Ambil nama kolom berdasarkan tipe"""
    numeric = [col for col, t in type_mapping.items() if t == 'Angka']
    date = [col for col, t in type_mapping.items() if t == 'Tanggal']
    category = [col for col, t in type_mapping.items() if t == 'Teks/Kategori']
    return numeric, date, category

# ==================== NEW: EXPORT FUNCTIONS ====================
def create_excel_with_formulas(analysis_result, analysis_type):
    """Buat file Excel dengan formula yang bisa dihitung ulang"""
    wb = Workbook()
    # JANGAN hapus sheet default dulu, akan dihapus di akhir
    
    # Style definitions
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if analysis_type == 'correlation':
        # Sheet 1: Ringkasan
        ws1 = wb.active
        ws1.title = 'Ringkasan'
        ws1.append(['Analisis Korelasi'])
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.append([])
        ws1.append(['Variabel X', analysis_result.get('x_var', '')])
        ws1.append(['Variabel Y', analysis_result.get('y_var', '')])
        ws1.append([])
        
        r = analysis_result['result']
        ws1.append(['Metrik', 'Nilai'])
        ws1['A6'].font = header_font
        ws1['A6'].fill = header_fill
        ws1['B6'].font = header_font
        ws1['B6'].fill = header_fill
        
        ws1.append(['Koefisien Korelasi (r)', r['r']])
        ws1.append(['P-value', r['p_value']])
        ws1.append(['Jumlah Data', r['n']])
        ws1.append(['Metode', r['method']])
        ws1.append(['Kekuatan', r['strength']])
        ws1.append(['Arah', r['direction']])
        ws1.append(['Signifikansi', r['significance']])
        
        # Sheet 2: Data Detail dengan Formula
        ws2 = wb.create_sheet('Data_Detail')
        ws2.append([analysis_result.get('x_var', 'X'), analysis_result.get('y_var', 'Y')])
        ws2['A1'].font = header_font
        ws2['A1'].fill = header_fill
        ws2['B1'].font = header_font
        ws2['B1'].fill = header_fill
        
        for x_val, y_val in zip(r['x'], r['y']):
            ws2.append([float(x_val), float(y_val)])
        
        # Tambahkan statistik dengan formula
        last_row = len(r['x']) + 1
        ws2.append([])
        ws2.append(['Statistik', 'Variabel X', 'Variabel Y'])
        stat_row = last_row + 2
        ws2[f'A{stat_row}'].font = header_font
        ws2[f'A{stat_row}'].fill = header_fill
        ws2[f'B{stat_row}'].font = header_font
        ws2[f'B{stat_row}'].fill = header_fill
        ws2[f'C{stat_row}'].font = header_font
        ws2[f'C{stat_row}'].fill = header_fill
        
        ws2.append(['Rata-rata', f'=AVERAGE(A2:A{last_row})', f'=AVERAGE(B2:B{last_row})'])
        ws2.append(['Median', f'=MEDIAN(A2:A{last_row})', f'=MEDIAN(B2:B{last_row})'])
        ws2.append(['Std Dev', f'=STDEV.S(A2:A{last_row})', f'=STDEV.S(B2:B{last_row})'])
        ws2.append(['Min', f'=MIN(A2:A{last_row})', f'=MIN(B2:B{last_row})'])
        ws2.append(['Max', f'=MAX(A2:A{last_row})', f'=MAX(B2:B{last_row})'])
        
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
    
    elif analysis_type == 'comparison':
        # Sheet 1: Ringkasan
        ws1 = wb.active
        ws1.title = 'Ringkasan'
        ws1.append(['Analisis Perbandingan Kelompok'])
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.append([])
        ws1.append(['Variabel Kelompok', analysis_result.get('group_var', '')])
        ws1.append(['Variabel Nilai', analysis_result.get('value_var', '')])
        ws1.append([])
        
        r = analysis_result['result']
        ws1.append(['Metrik', 'Nilai'])
        ws1['A6'].font = header_font
        ws1['A6'].fill = header_fill
        ws1['B6'].font = header_font
        ws1['B6'].fill = header_fill
        
        ws1.append(['Metode Uji', r['test_name']])
        ws1.append(['Statistik Uji', r['stat']])
        ws1.append(['P-value', r['p_value']])
        ws1.append(['Jumlah Kelompok', r['n_groups']])
        ws1.append(['Signifikansi', r['significance']])
        
        # Sheet 2: Statistik per Kelompok
        ws2 = wb.create_sheet('Statistik_Kelompok')
        for r_idx in dataframe_to_rows(r['summary'], index=False, header=True):
            ws2.append(r_idx)
        
        # Format header
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in ws2.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column[0].column_letter].width = adjusted_width
    
    elif analysis_type == 'trend':
        # Sheet 1: Ringkasan
        ws1 = wb.active
        ws1.title = 'Ringkasan'
        ws1.append(['Analisis Tren Waktu'])
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.append([])
        ws1.append(['Variabel Waktu', analysis_result.get('date_var', '')])
        ws1.append(['Variabel Nilai', analysis_result.get('value_var', '')])
        ws1.append(['Periode', analysis_result.get('freq', '')])
        ws1.append([])
        
        r = analysis_result['result']
        ws1.append(['Metrik', 'Nilai'])
        ws1['A7'].font = header_font
        ws1['A7'].fill = header_fill
        ws1['B7'].font = header_font
        ws1['B7'].fill = header_fill
        
        ws1.append(['Korelasi Tren (r)', r['r']])
        ws1.append(['P-value', r['p_value']])
        ws1.append(['Deskripsi Tren', r['trend_desc']])
        ws1.append(['Kekuatan Tren', r['trend_strength']])
        ws1.append(['Nilai Awal', r['first_val']])
        ws1.append(['Nilai Akhir', r['last_val']])
        ws1.append(['Perubahan Absolut', r['change']])
        ws1.append(['Perubahan Persen', f"{r['change_pct']:.2f}%"])
        ws1.append(['Jumlah Periode', r['n_periods']])
        
        # Sheet 2: Data Time Series dengan Formula
        ws2 = wb.create_sheet('Data_Time_Series')
        ws2.append(['Tanggal', 'Nilai'])
        ws2['A1'].font = header_font
        ws2['A1'].fill = header_fill
        ws2['B1'].font = header_font
        ws2['B1'].fill = header_fill
        
        for date, val in zip(r['dates'], r['values']):
            ws2.append([date.strftime('%Y-%m-%d'), float(val)])
        
        # Tambahkan statistik dengan formula
        last_row = len(r['values']) + 1
        ws2.append([])
        ws2.append(['Statistik', 'Nilai'])
        stat_row = last_row + 2
        ws2[f'A{stat_row}'].font = header_font
        ws2[f'A{stat_row}'].fill = header_fill
        ws2[f'B{stat_row}'].font = header_font
        ws2[f'B{stat_row}'].fill = header_fill
        
        ws2.append(['Rata-rata', f'=AVERAGE(B2:B{last_row})'])
        ws2.append(['Median', f'=MEDIAN(B2:B{last_row})'])
        ws2.append(['Std Dev', f'=STDEV.S(B2:B{last_row})'])
        ws2.append(['Min', f'=MIN(B2:B{last_row})'])
        ws2.append(['Max', f'=MAX(B2:B{last_row})'])
        ws2.append(['Total', f'=SUM(B2:B{last_row})'])
        
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 15
    
    elif analysis_type == 'regression':
        # Sheet 1: Ringkasan
        ws1 = wb.active
        ws1.title = 'Ringkasan'
        ws1.append(['Analisis Regresi Linear'])
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.append([])
        ws1.append(['Variabel X', analysis_result.get('x_var', '')])
        ws1.append(['Variabel Y', analysis_result.get('y_var', '')])
        ws1.append([])
        
        r = analysis_result['result']
        ws1.append(['Metrik', 'Nilai'])
        ws1['A6'].font = header_font
        ws1['A6'].fill = header_fill
        ws1['B6'].font = header_font
        ws1['B6'].fill = header_fill
        
        ws1.append(['R²', r['r_squared']])
        ws1.append(['Slope (b)', r['slope']])
        ws1.append(['Intercept (a)', r['intercept']])
        ws1.append(['P-value', r['p_value']])
        ws1.append(['RMSE', r['rmse']])
        ws1.append(['MAE', r['mae']])
        
        # Sheet 2: Data Detail
        ws2 = wb.create_sheet('Data_Detail')
        ws2.append([analysis_result.get('x_var', 'X'), analysis_result.get('y_var', 'Y'), 'Y_Prediksi', 'Residual'])
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        for x_val, y_val, y_pred, resid in zip(r['x'], r['y'], r['y_pred'], r['residuals']):
            ws2.append([float(x_val), float(y_val), float(y_pred), float(resid)])
        
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        ws2.column_dimensions['D'].width = 15
    
    elif analysis_type == 'outlier':
        # Sheet 1: Ringkasan
        ws1 = wb.active
        ws1.title = 'Ringkasan'
        ws1.append(['Deteksi Outlier'])
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.append([])
        ws1.append(['Variabel', analysis_result.get('var_col', '')])
        ws1.append([])
        
        r = analysis_result['result']
        ws1.append(['Metrik', 'Nilai'])
        ws1['A5'].font = header_font
        ws1['A5'].fill = header_fill
        ws1['B5'].font = header_font
        ws1['B5'].fill = header_fill
        
        ws1.append(['Total Data', r['n_total']])
        ws1.append(['Outlier (IQR)', r['n_outliers_iqr']])
        ws1.append(['% Outlier (IQR)', f"{r['pct_outliers_iqr']:.2f}%"])
        ws1.append(['Lower Bound (IQR)', r['lower_bound']])
        ws1.append(['Upper Bound (IQR)', r['upper_bound']])
        ws1.append(['Outlier (Z-score)', r['n_outliers_zscore']])
        ws1.append(['% Outlier (Z-score)', f"{r['pct_outliers_zscore']:.2f}%"])
        
        # Sheet 2: Data Lengkap
        ws2 = wb.create_sheet('All_Data')
        ws2.append(['Nilai'])
        ws2['A1'].font = header_font
        ws2['A1'].fill = header_fill
        
        for val in r['all_values']:
            ws2.append([float(val)])
    
    else:
        # Default untuk analysis type lain
        ws1 = wb.active
        ws1.title = 'Hasil_Analisis'
        ws1.append(['Hasil Analisis'])
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.append([])
        ws1.append(['Tipe Analisis', analysis_type])
        ws1.append(['Timestamp', analysis_result.get('timestamp', 'N/A')])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
    
def export_to_csv(analysis_result, analysis_type):
    """Export hasil analisis ke CSV"""
    if analysis_type == 'correlation':
        r = analysis_result['result']
        
        # CSV 1: Ringkasan
        summary_data = {
            'Metrik': ['Koefisien Korelasi (r)', 'P-value', 'Jumlah Data', 'Metode', 
                       'Kekuatan', 'Arah', 'Signifikansi'],
            'Nilai': [r['r'], r['p_value'], r['n'], r['method'], 
                     r['strength'], r['direction'], r['significance']]
        }
        df_summary = pd.DataFrame(summary_data)
        
        # CSV 2: Data Detail
        df_detail = pd.DataFrame({
            analysis_result.get('x_var', 'X'): r['x'],
            analysis_result.get('y_var', 'Y'): r['y']
        })
        
        # CSV 3: Statistik
        df_stats = pd.DataFrame({
            'Statistik': ['Rata-rata', 'Median', 'Std Dev', 'Min', 'Max'],
            'Variabel X': [r['stats']['mean_x'], np.median(r['x']), r['stats']['std_x'], 
                          np.min(r['x']), np.max(r['x'])],
            'Variabel Y': [r['stats']['mean_y'], np.median(r['y']), r['stats']['std_y'],
                          np.min(r['y']), np.max(r['y'])]
        })
        
        return df_summary, df_detail, df_stats
    
    elif analysis_type == 'comparison':
        r = analysis_result['result']
        
        # CSV 1: Ringkasan
        summary_data = {
            'Metrik': ['Metode Uji', 'Statistik Uji', 'P-value', 'Jumlah Kelompok', 'Signifikansi'],
            'Nilai': [r['test_name'], r['stat'], r['p_value'], r['n_groups'], r['significance']]
        }
        df_summary = pd.DataFrame(summary_data)
        
        # CSV 2: Statistik per Kelompok
        df_detail = r['summary']
        
        return df_summary, df_detail, None
    
    elif analysis_type == 'trend':
        r = analysis_result['result']
        
        # CSV 1: Ringkasan
        summary_data = {
            'Metrik': ['Korelasi Tren (r)', 'P-value', 'Deskripsi Tren', 'Kekuatan Tren',
                      'Nilai Awal', 'Nilai Akhir', 'Perubahan Absolut', 'Perubahan Persen', 'Jumlah Periode'],
            'Nilai': [r['r'], r['p_value'], r['trend_desc'], r['trend_strength'],
                     r['first_val'], r['last_val'], r['change'], f"{r['change_pct']:.2f}%", r['n_periods']]
        }
        df_summary = pd.DataFrame(summary_data)
        
        # CSV 2: Data Time Series
        df_detail = pd.DataFrame({
            'Tanggal': [d.strftime('%Y-%m-%d') for d in r['dates']],
            'Nilai': r['values']
        })
        
        # CSV 3: Statistik
        df_stats = pd.DataFrame({
            'Statistik': ['Rata-rata', 'Median', 'Std Dev', 'Min', 'Max', 'Total'],
            'Nilai': [r['mean'], np.median(r['values']), r['std'], 
                     np.min(r['values']), np.max(r['values']), np.sum(r['values'])]
        })
        
        return df_summary, df_detail, df_stats
    
    return None, None, None

# ==================== ANALYSIS FUNCTIONS ====================
def analyze_correlation(df, x_col, y_col):
    """Analisis korelasi dengan uji asumsi"""
    data_clean = df[[x_col, y_col]].dropna()
    x = data_clean[x_col].values
    y = data_clean[y_col].values
    n = len(x)
    
    if n < 8:
        return {'error': 'Data terlalu sedikit (minimal 8 pasangan data)'}
    
    _, p_norm_x = stats.shapiro(x[:min(n, 5000)])
    _, p_norm_y = stats.shapiro(y[:min(n, 5000)])
    is_normal = (p_norm_x > 0.05) and (p_norm_y > 0.05)
    
    if is_normal:
        r, p_val = stats.pearsonr(x, y)
        method = 'Pearson'
        method_explain = 'Melihat apakah dua angka cenderung naik/turun bersama secara cukup lurus.'
    else:
        r, p_val = stats.spearmanr(x, y)
        method = 'Spearman'
        method_explain = 'Melihat kecenderungan naik/turun bersama walau polanya tidak harus lurus; lebih tahan terhadap nilai yang terlalu ekstrem.'

    abs_r = abs(r)
    if abs_r > 0.7:
        strength = 'sangat kuat'
        practical = 'Hubungan ini kuat, jadi bisa dipakai sebagai acuan (tetap cek faktor lain).'
    elif abs_r > 0.5:
        strength = 'kuat'
        practical = 'Ada kecenderungan yang cukup jelas, layak dipertimbangkan.'
    elif abs_r > 0.3:
        strength = 'sedang'
        practical = 'Ada tanda hubungan, tapi pengaruhnya kecil. Faktor lain masih penting.'
    else:
        strength = 'lemah atau tidak ada'
        practical = 'Hampir tidak ada kaitan yang jelas antara keduanya.'

    direction = 'positif' if r > 0 else 'negatif'
    significance = 'signifikan' if p_val < 0.05 else 'tidak signifikan'
    
    return {
        'x': x, 'y': y, 'n': n,
        'method': method,
        'method_explain': method_explain,
        'r': r, 'p_value': p_val,
        'is_normal': is_normal,
        'p_norm_x': p_norm_x,
        'p_norm_y': p_norm_y,
        'strength': strength,
        'direction': direction,
        'significance': significance,
        'practical': practical,
        'stats': {
            'mean_x': x.mean(),
            'mean_y': y.mean(),
            'std_x': x.std(),
            'std_y': y.std()
        }
    }

def analyze_group_comparison(df, group_col, value_col):
    """Perbandingan kelompok dengan uji statistik"""
    data_clean = df[[group_col, value_col]].dropna()
    
    top_groups = data_clean[group_col].value_counts().head(10).index
    data_clean = data_clean[data_clean[group_col].isin(top_groups)]
    
    summary = data_clean.groupby(group_col)[value_col].agg([
        ('Jumlah Data', 'count'),
        ('Total', 'sum'),
        ('Rata-rata', 'mean'),
        ('Median', 'median'),
        ('Std Dev', 'std'),
        ('Min', 'min'),
        ('Max', 'max')
    ]).round(2).reset_index()
    
    groups = [data_clean[data_clean[group_col] == g][value_col].values 
              for g in top_groups]
    
    normal_flags = []
    for arr in groups:
        if len(arr) >= 3:
            _, p = stats.shapiro(arr[:min(len(arr), 5000)])
            normal_flags.append(p > 0.05)
    is_normal = all(normal_flags) if normal_flags else False
    
    if len(groups) == 2:
        if is_normal:
            stat, p_val = stats.ttest_ind(groups[0], groups[1], equal_var=False)
            test_name = 'Uji beda rata-rata 2 kelompok'
            test_explain = 'Mengecek apakah rata-rata 2 kelompok benar-benar berbeda.'
        else:
            stat, p_val = stats.mannwhitneyu(groups[0], groups[1])
            test_name = 'Uji beda nilai tengah 2 kelompok'
            test_explain = 'Mengecek apakah nilai tengah 2 kelompok berbeda (tanpa syarat data harus ‘rapi/normal’).'
    else:
        if is_normal:
            stat, p_val = stats.f_oneway(*groups)
            test_name = 'Uji beda rata-rata beberapa kelompok'
            test_explain = 'Mengecek apakah rata-rata beberapa kelompok ada yang berbeda.'
        else:
            stat, p_val = stats.kruskal(*groups)
            test_name = 'Uji beda nilai tengah beberapa kelompok'
            test_explain = 'Mengecek apakah nilai tengah beberapa kelompok ada yang berbeda (tanpa syarat data harus ‘rapi/normal’).'

    significance = 'signifikan' if p_val < 0.05 else 'tidak signifikan'
    
    return {
        'summary': summary,
        'test_name': test_name,
        'test_explain': test_explain,
        'stat': stat,
        'p_value': p_val,
        'is_normal': is_normal,
        'significance': significance,
        'n_groups': len(groups)
    }

def analyze_trend(df, date_col, value_col, freq='M'):
    """Analisis tren waktu"""
    data_clean = df[[date_col, value_col]].dropna().copy()
    data_clean[date_col] = pd.to_datetime(data_clean[date_col])
    data_clean = data_clean.sort_values(date_col)
    
    ts = data_clean.set_index(date_col)[value_col].resample(freq).sum(min_count=1)
    ts_clean = ts.dropna()
    
    if len(ts_clean) < 3:
        return {'error': 'Data terlalu sedikit untuk analisis tren'}
    
    x_time = np.arange(len(ts_clean))
    r, p_val = stats.spearmanr(x_time, ts_clean.values)
    
    first_val = ts_clean.iloc[0]
    last_val = ts_clean.iloc[-1]
    change = last_val - first_val
    change_pct = (change / first_val * 100) if first_val != 0 else 0
    
    if r > 0.3 and p_val < 0.05:
        trend_desc = 'naik'
        trend_strength = 'kuat' if abs(r) > 0.7 else 'sedang'
    elif r < -0.3 and p_val < 0.05:
        trend_desc = 'turun'
        trend_strength = 'kuat' if abs(r) > 0.7 else 'sedang'
    else:
        trend_desc = 'stabil/fluktuatif'
        trend_strength = 'lemah'
    
    return {
        'ts': ts_clean,
        'dates': ts_clean.index,
        'values': ts_clean.values,
        'r': r,
        'p_value': p_val,
        'trend_desc': trend_desc,
        'trend_strength': trend_strength,
        'first_val': first_val,
        'last_val': last_val,
        'change': change,
        'change_pct': change_pct,
        'mean': ts_clean.mean(),
        'std': ts_clean.std(),
        'n_periods': len(ts_clean)
    }

def analyze_distribution(df, var_col):
    """Analisis distribusi variabel numerik"""
    values = df[var_col].dropna().values
    
    if len(values) < 3:
        return {'error': 'Data terlalu sedikit'}
    
    mean_val = values.mean()
    median_val = np.median(values)
    std_val = values.std()
    min_val = values.min()
    max_val = values.max()
    q1 = np.percentile(values, 25)
    q3 = np.percentile(values, 75)
    iqr = q3 - q1
    cv = (std_val / mean_val * 100) if mean_val != 0 else 0
    
    _, p_norm = stats.shapiro(values[:min(len(values), 5000)])
    is_normal = p_norm > 0.05
    
    if mean_val > median_val + std_val * 0.5:
        skewness = 'condong ke nilai besar'
        skew_detail = 'Ada beberapa angka yang jauh lebih besar dari kebanyakan.'
    elif mean_val < median_val - std_val * 0.5:
        skewness = 'condong ke nilai kecil'
        skew_detail = 'Ada beberapa angka yang jauh lebih kecil dari kebanyakan.'
    else:
        skewness = 'cukup seimbang'
        skew_detail = 'Sebaran nilai cukup merata di sekitar nilai tengah.'

    if cv < 20:
        variability = 'nilainya cenderung mirip-mirip'
    elif cv < 40:
        variability = 'nilainya cukup mirip'
    else:
        variability = 'nilainya sangat bervariasi'
    
    return {
        'values': values,
        'n': len(values),
        'mean': mean_val,
        'median': median_val,
        'std': std_val,
        'min': min_val,
        'max': max_val,
        'q1': q1,
        'q3': q3,
        'iqr': iqr,
        'cv': cv,
        'range': max_val - min_val,
        'is_normal': is_normal,
        'p_norm': p_norm,
        'skewness': skewness,
        'skew_detail': skew_detail,
        'variability': variability
    }

# ==================== NEW: ADVANCED ANALYSIS FUNCTIONS ====================
def analyze_regression(df, x_col, y_col):
    """Analisis regresi linear sederhana"""
    data_clean = df[[x_col, y_col]].dropna()
    x = data_clean[x_col].values
    y = data_clean[y_col].values
    
    if len(x) < 3:
        return {'error': 'Data terlalu sedikit untuk regresi'}
    
    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
    
    # Predict values
    y_pred = slope * x + intercept
    residuals = y - y_pred
    
    # R-squared
    r_squared = r_value ** 2
    
    # RMSE
    rmse = np.sqrt(np.mean(residuals ** 2))
    
    # MAE
    mae = np.mean(np.abs(residuals))
    
    return {
        'slope': slope,
        'intercept': intercept,
        'r_value': r_value,
        'r_squared': r_squared,
        'p_value': p_value,
        'std_err': std_err,
        'rmse': rmse,
        'mae': mae,
        'x': x,
        'y': y,
        'y_pred': y_pred,
        'residuals': residuals,
        'n': len(x)
    }

def analyze_outliers(df, var_col):
    """Deteksi outlier menggunakan metode IQR dan Z-score"""
    values = df[var_col].dropna().values
    
    if len(values) < 3:
        return {'error': 'Data terlalu sedikit'}
    
    # IQR method
    q1 = np.percentile(values, 25)
    q3 = np.percentile(values, 75)
    iqr = q3 - q1
    lower_bound = q1 - 1.5 * iqr
    upper_bound = q3 + 1.5 * iqr
    
    outliers_iqr = values[(values < lower_bound) | (values > upper_bound)]
    outlier_indices_iqr = np.where((values < lower_bound) | (values > upper_bound))[0]
    
    # Z-score method
    mean = values.mean()
    std = values.std()
    z_scores = np.abs((values - mean) / std)
    outliers_zscore = values[z_scores > 3]
    outlier_indices_zscore = np.where(z_scores > 3)[0]
    
    return {
        'n_total': len(values),
        'outliers_iqr': outliers_iqr,
        'n_outliers_iqr': len(outliers_iqr),
        'pct_outliers_iqr': len(outliers_iqr) / len(values) * 100,
        'lower_bound': lower_bound,
        'upper_bound': upper_bound,
        'outliers_zscore': outliers_zscore,
        'n_outliers_zscore': len(outliers_zscore),
        'pct_outliers_zscore': len(outliers_zscore) / len(values) * 100,
        'all_values': values
    }

def analyze_multiple_regression(df, x_cols, y_col):
    """Regresi berganda - prediksi Y dari multiple X"""
    from sklearn.linear_model import LinearRegression
    from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
    
    cols_to_use = x_cols + [y_col]
    data_clean = df[cols_to_use].dropna()
    
    if len(data_clean) < len(x_cols) + 2:
        return {'error': f'Data terlalu sedikit (minimal {len(x_cols) + 2} baris)'}
    
    X = data_clean[x_cols].values
    y = data_clean[y_col].values
    
    # Fit model
    model = LinearRegression()
    model.fit(X, y)
    
    # Predictions
    y_pred = model.predict(X)
    residuals = y - y_pred
    
    # Metrics
    r2 = r2_score(y, y_pred)
    rmse = np.sqrt(mean_squared_error(y, y_pred))
    mae = mean_absolute_error(y, y_pred)
    
    # Adjusted R²
    n = len(y)
    p = len(x_cols)
    adj_r2 = 1 - (1 - r2) * (n - 1) / (n - p - 1)
    
    # Coefficients with names
    coefs = dict(zip(x_cols, model.coef_))
    
    return {
        'intercept': model.intercept_,
        'coefficients': coefs,
        'r2': r2,
        'adj_r2': adj_r2,
        'rmse': rmse,
        'mae': mae,
        'n': n,
        'n_predictors': p,
        'y_actual': y,
        'y_pred': y_pred,
        'residuals': residuals,
        'x_cols': x_cols
    }

def analyze_chi_square(df, var1, var2):
    """Uji Chi-Square untuk hubungan antar variabel kategorikal"""
    from scipy.stats import chi2_contingency
    
    data_clean = df[[var1, var2]].dropna()
    
    if len(data_clean) < 5:
        return {'error': 'Data terlalu sedikit (minimal 5 observasi)'}
    
    # Create contingency table
    contingency_table = pd.crosstab(data_clean[var1], data_clean[var2])
    
    # Chi-square test
    chi2, p_value, dof, expected = chi2_contingency(contingency_table)
    
    # Cramér's V (effect size)
    n = contingency_table.sum().sum()
    min_dim = min(contingency_table.shape[0] - 1, contingency_table.shape[1] - 1)
    cramers_v = np.sqrt(chi2 / (n * min_dim))
    
    # Interpretation
    if cramers_v < 0.1:
        strength = 'sangat lemah'
    elif cramers_v < 0.3:
        strength = 'lemah'
    elif cramers_v < 0.5:
        strength = 'sedang'
    else:
        strength = 'kuat'
    
    significance = 'signifikan' if p_value < 0.05 else 'tidak signifikan'
    
    return {
        'contingency_table': contingency_table,
        'chi2': chi2,
        'p_value': p_value,
        'dof': dof,
        'cramers_v': cramers_v,
        'strength': strength,
        'significance': significance,
        'n': n
    }

def analyze_time_series_forecast(df, date_col, value_col, periods=3):
    """Simple time series forecasting menggunakan linear trend"""
    data_clean = df[[date_col, value_col]].dropna().copy()
    data_clean[date_col] = pd.to_datetime(data_clean[date_col])
    data_clean = data_clean.sort_values(date_col)
    
    if len(data_clean) < 3:
        return {'error': 'Data terlalu sedikit untuk forecasting'}
    
    # Prepare data
    data_clean['time_index'] = np.arange(len(data_clean))
    
    # Linear regression for trend
    X = data_clean['time_index'].values.reshape(-1, 1)
    y = data_clean[value_col].values
    
    slope, intercept, r_value, p_value, std_err = stats.linregress(
        data_clean['time_index'], y
    )
    
    # Forecast
    last_time_index = data_clean['time_index'].iloc[-1]
    future_indices = np.arange(last_time_index + 1, last_time_index + periods + 1)
    forecast_values = slope * future_indices + intercept
    
    # Generate future dates
    last_date = data_clean[date_col].iloc[-1]
    date_diff = data_clean[date_col].diff().median()
    future_dates = [last_date + date_diff * (i + 1) for i in range(periods)]
    
    # Historical fit
    fitted_values = slope * data_clean['time_index'] + intercept
    residuals = y - fitted_values
    rmse = np.sqrt(np.mean(residuals ** 2))
    
    return {
        'historical_dates': data_clean[date_col].values,
        'historical_values': y,
        'fitted_values': fitted_values,
        'future_dates': future_dates,
        'forecast_values': forecast_values,
        'slope': slope,
        'intercept': intercept,
        'r_squared': r_value ** 2,
        'rmse': rmse,
        'n_historical': len(y),
        'n_forecast': periods
    }

def analyze_correlation_matrix(df, num_cols):
    """Analisis korelasi untuk multiple variabel sekaligus"""
    if len(num_cols) < 2:
        return {'error': 'Minimal 2 variabel numerik diperlukan'}
    
    data_clean = df[num_cols].dropna()
    
    if len(data_clean) < 3:
        return {'error': 'Data terlalu sedikit setelah cleaning'}
    
    # Correlation matrix
    corr_matrix = data_clean.corr()
    
    # Find strongest correlations
    corr_pairs = []
    for i in range(len(num_cols)):
        for j in range(i + 1, len(num_cols)):
            corr_pairs.append({
                'var1': num_cols[i],
                'var2': num_cols[j],
                'correlation': corr_matrix.iloc[i, j]
            })
    
    corr_df = pd.DataFrame(corr_pairs)
    corr_df = corr_df.sort_values('correlation', key=abs, ascending=False)
    
    return {
        'corr_matrix': corr_matrix,
        'corr_pairs': corr_df,
        'n': len(data_clean),
        'variables': num_cols
    }

def analyze_normality_test(df, var_col):
    """Uji normalitas lengkap dengan visualisasi"""
    values = df[var_col].dropna().values
    
    if len(values) < 3:
        return {'error': 'Data terlalu sedikit'}
    
    # Shapiro-Wilk test
    stat_sw, p_sw = stats.shapiro(values[:min(len(values), 5000)])
    
    # Kolmogorov-Smirnov test
    stat_ks, p_ks = stats.kstest(values, 'norm', args=(values.mean(), values.std()))
    
    # Anderson-Darling test
    result_ad = stats.anderson(values, dist='norm')
    
    # Skewness and Kurtosis
    skewness = stats.skew(values)
    kurtosis = stats.kurtosis(values)
    
    # Interpretation
    is_normal_sw = p_sw > 0.05
    is_normal_ks = p_ks > 0.05
    
    if is_normal_sw and is_normal_ks:
        conclusion = 'Data mengikuti distribusi normal'
    elif is_normal_sw or is_normal_ks:
        conclusion = 'Data mendekati distribusi normal (hasil uji berbeda)'
    else:
        conclusion = 'Data tidak mengikuti distribusi normal'
    
    return {
        'values': values,
        'n': len(values),
        'shapiro_stat': stat_sw,
        'shapiro_p': p_sw,
        'ks_stat': stat_ks,
        'ks_p': p_ks,
        'anderson_stat': result_ad.statistic,
        'anderson_critical': result_ad.critical_values[2],  # 5% level
        'skewness': skewness,
        'kurtosis': kurtosis,
        'is_normal_sw': is_normal_sw,
        'is_normal_ks': is_normal_ks,
        'conclusion': conclusion
    }

def analyze_variance_homogeneity(df, group_col, value_col):
    """Uji homogenitas varians (Levene's test)"""
    data_clean = df[[group_col, value_col]].dropna()
    
    top_groups = data_clean[group_col].value_counts().head(10).index
    data_clean = data_clean[data_clean[group_col].isin(top_groups)]
    
    groups = [data_clean[data_clean[group_col] == g][value_col].values 
              for g in top_groups]
    
    # Filter out groups with less than 2 observations
    groups = [g for g in groups if len(g) >= 2]
    
    if len(groups) < 2:
        return {'error': 'Tidak cukup kelompok dengan data yang memadai'}
    
    # Levene's test
    stat, p_value = stats.levene(*groups)
    
    is_homogeneous = p_value > 0.05
    
    conclusion = 'Varians antar kelompok homogen (sama)' if is_homogeneous else 'Varians antar kelompok tidak homogen'
    
    return {
        'stat': stat,
        'p_value': p_value,
        'is_homogeneous': is_homogeneous,
        'conclusion': conclusion,
        'n_groups': len(groups),
        'recommendation': 'Gunakan t-test/ANOVA standar' if is_homogeneous else 'Gunakan Welch t-test atau transformasi data'
    }

# ==================== NAVIGATION ====================
st.sidebar.title("Dashboard Analisis PSDKP")
st.sidebar.markdown("---")
page = st.sidebar.radio(
    "Navigasi",
    ["Intro", "Input Data", "Statistik Deskriptif", "Analisis", "Hasil & Export", "Kesimpulan"],
    key="navigation"
)

# ==================== PAGE 1: INTRO ====================
if page == "Intro":
    st.title("Dashboard Analisis PSDKP")
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        ### Fitur Utama
        
        **1. Analisis Statistik Lengkap**
        - Korelasi (Pearson/Spearman)
        - Perbandingan Kelompok (t-test, ANOVA, dll)
        - Analisis Tren Waktu
        - Regresi Linear
        - Deteksi Outlier (IQR & Z-score)
        
        **2. Export Hasil Analisis**
        - Download CSV hasil perhitungan
        - File bisa dihitung ulang di Excel
        
        **3. Visualisasi Interaktif**
        - Grafik yang bisa di-zoom dan di-explore
        - Tabel hasil yang rapi dan lengkap
        - Interpretasi dalam Bahasa Indonesia
        
        **4. Uji Statistik Otomatis**
        - Sistem pilih metode yang tepat
        - Uji asumsi dijelaskan dengan jelas
        - P-value dan significance level
        """)
    
    with col2:
        st.markdown("""
        ### Cara Kerja
        
        **Step 1: Upload Data**
        - Upload file Excel (.xlsx) atau CSV
        - Sistem deteksi otomatis tipe kolom
        - Konfirmasi & sesuaikan jika perlu
        
        **Step 2: Pilih Analisis**
        - Pilih jenis analisis sesuai kebutuhan
        - Pilih variabel yang ingin dianalisis
        - Sistem otomatis pilih metode statistik
        
        **Step 3: Lihat Hasil**
        - Visualisasi interaktif
        - Hasil uji statistik dengan interpretasi
        - Uji asumsi dijelaskan dengan sederhana
        
        **Step 4: Download Hasil**
        - Download CSV hasil perhitungan
        - Simpan untuk dokumentasi/laporan
        
        ### Fitur
        
        - Export CSV hasil analisis
        - Analisis regresi linear
        - Deteksi outlier otomatis
        - Riwayat analisis tersimpan
        - Perhitungan langsung di Excel
        """)
    
    st.markdown("""
    <div class="success-box">
        <strong>Mulai Sekarang!</strong><br>
        Klik "Input Data" di menu samping untuk upload file dan mulai analisis.
    </div>
    """, unsafe_allow_html=True)

# ==================== PAGE 2: INPUT DATA ====================
elif page == "Input Data":
    st.header("Input Data")
    
    st.markdown("""
    <div class="info-box">
        <strong>Format yang Didukung:</strong> Excel (.xlsx, .xls), CSV (.csv)<br>
        <strong>Syarat:</strong> Baris pertama harus header/nama kolom, minimal 10 baris data
    </div>
    """, unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader(
        "Upload File Data Anda",
        type=['xlsx', 'xls', 'csv'],
        help="Pilih file Excel atau CSV"
    )
    
    if uploaded_file is not None:
        df, error = read_file(uploaded_file)
        
        if error:
            st.error(f"Error: {error}")
        else:
            st.session_state.df_raw = df
            
            st.success(f"File berhasil dibaca: {len(df):,} baris × {len(df.columns)} kolom")
            
            st.subheader("Preview Data")
            st.dataframe(df.head(20), use_container_width=True)
            st.caption(f"Menampilkan 20 dari {len(df):,} baris")
            
            st.subheader("Konfirmasi Tipe Data Kolom")
            st.markdown("""
            <div class="info-box">
                Sistem sudah mendeteksi tipe data otomatis. Periksa dan ubah jika ada yang salah.
            </div>
            """, unsafe_allow_html=True)
            
            type_options = ['Angka', 'Teks/Kategori', 'Tanggal', 'Boolean']
            column_types = {}
            
            cols_per_row = 3
            for i in range(0, len(df.columns), cols_per_row):
                cols = st.columns(cols_per_row)
                for j, col in enumerate(df.columns[i:i+cols_per_row]):
                    with cols[j]:
                        inferred = infer_column_type(df[col])
                        st.markdown(f"**{col}**")
                        st.caption(f"Contoh: {df[col].dropna().iloc[0] if len(df[col].dropna()) > 0 else 'N/A'}")
                        
                        column_types[col] = st.selectbox(
                            "Tipe",
                            type_options,
                            index=type_options.index(inferred) if inferred in type_options else 1,
                            key=f"type_{col}",
                            label_visibility="collapsed"
                        )
            
            st.session_state.column_types = column_types
            
            if st.button("Simpan & Terapkan Tipe Data", type="primary", use_container_width=True):
                df_processed, warnings = apply_column_types(df, column_types)
                st.session_state.df = df_processed
                
                st.success("Tipe data berhasil diterapkan")
                
                if warnings:
                    with st.expander("Peringatan Konversi", expanded=True):
                        for w in warnings:
                            st.warning(w)
                
                num_cols, date_cols, cat_cols = get_columns_by_type(df_processed, column_types)
                
                col1, col2, col3 = st.columns(3)
                col1.metric("Kolom Angka", len(num_cols))
                col2.metric("Kolom Tanggal", len(date_cols))
                col3.metric("Kolom Kategori", len(cat_cols))
                
                st.markdown("""
                <div class="success-box">
                    Data siap dianalisis! Lanjut ke "Analisis" untuk memilih jenis analisis.
                </div>
                """, unsafe_allow_html=True)
    else:
        st.info("Belum ada file yang diupload. Pilih file untuk memulai.")

# ==================== PAGE: STATISTIK DESKRIPTIF ====================
elif page == "Statistik Deskriptif":
    st.header("Statistik Deskriptif")
    
    if st.session_state.df is None:
        st.warning("Belum ada data. Upload dan terapkan tipe data di 'Input Data' dulu.")
    else:
        df = st.session_state.df
        types = st.session_state.column_types
        num_cols, date_cols, cat_cols = get_columns_by_type(df, types)
        
        st.markdown("""
        <div class="info-box">
            <strong>Tahap Eksplorasi Data</strong><br>
            Pahami karakteristik data sebelum melakukan analisis lanjutan.
            Statistik deskriptif membantu mendeteksi pola, anomali, dan karakteristik dasar dari data.
        </div>
        """, unsafe_allow_html=True)
        
        # ==================== OVERVIEW DATA ====================
        st.subheader("Overview Dataset")
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Baris", f"{len(df):,}")
        col2.metric("Total Kolom", f"{len(df.columns):,}")
        col3.metric("Kolom Numerik", f"{len(num_cols)}")
        col4.metric("Kolom Kategorikal", f"{len(cat_cols)}")
        
        # Missing values overview
        st.markdown("### Data Kosong")
        missing_df = pd.DataFrame({
            'Kolom': df.columns,
            'Missing Count': df.isnull().sum().values,
            'Missing %': (df.isnull().sum().values / len(df) * 100).round(2)
        })
        missing_df = missing_df[missing_df['Missing Count'] > 0].sort_values('Missing Count', ascending=False)
        
        if len(missing_df) > 0:
            st.dataframe(missing_df, use_container_width=True)
            st.warning(f"Ditemukan {len(missing_df)} kolom dengan data kosong")
        else:
            st.success("Tidak ada data kosong dalam dataset")
        
        # ==================== STATISTIK NUMERIK ====================
        if len(num_cols) > 0:
            st.markdown("---")
            st.subheader("Statistik Variabel Numerik")
            
            # Pilih variabel untuk analisis detail
            selected_num = st.selectbox("Pilih variabel numerik untuk analisis detail:", num_cols)
            
            if selected_num:
                values = df[selected_num].dropna()
                
                # Metrics
                col1, col2, col3, col4, col5 = st.columns(5)
                col1.metric("Rata-rata", f"{values.mean():.2f}")
                col2.metric("Median", f"{values.median():.2f}")
                col3.metric("Std Dev", f"{values.std():.2f}")
                col4.metric("Min", f"{values.min():.2f}")
                col5.metric("Max", f"{values.max():.2f}")
                
                col1, col2, col3, col4, col5 = st.columns(5)
                q1 = values.quantile(0.25)
                q3 = values.quantile(0.75)
                iqr = q3 - q1
                col1.metric("Q1 (25%)", f"{q1:.2f}")
                col2.metric("Q3 (75%)", f"{q3:.2f}")
                col3.metric("IQR", f"{iqr:.2f}")
                col4.metric("Skewness", f"{stats.skew(values):.2f}")
                col5.metric("Kurtosis", f"{stats.kurtosis(values):.2f}")
                
                # Visualisasi
                tab1, tab2, tab3 = st.tabs(["Histogram", "Box Plot", "Distribution"])
                
                with tab1:
                    fig = px.histogram(
                        x=values,
                        nbins=30,
                        title=f"Distribusi {selected_num}",
                        labels={'x': selected_num, 'y': 'Frekuensi'},
                        marginal="box"
                    )
                    fig.update_traces(marker_color='#667eea')
                    st.plotly_chart(fig, use_container_width=True)
                
                with tab2:
                    fig = go.Figure()
                    fig.add_trace(go.Box(
                        y=values,
                        name=selected_num,
                        marker_color='#667eea',
                        boxmean='sd'
                    ))
                    fig.update_layout(
                        title=f"Box Plot: {selected_num}",
                        yaxis_title=selected_num
                    )
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # Interpretasi box plot
                    st.markdown("""
                    **Cara Membaca Box Plot:**
                    - Garis tengah kotak = Median (Q2)
                    - Batas bawah kotak = Q1 (25%)
                    - Batas atas kotak = Q3 (75%)
                    - Panjang kotak = IQR (Interquartile Range)
                    - Whiskers = Min/Max (dalam 1.5×IQR)
                    - Titik di luar whiskers = Outlier potensial
                    """)
                
                with tab3:
                    # QQ plot untuk uji normalitas visual
                    fig = go.Figure()
                    
                    # Theoretical quantiles
                    sorted_values = np.sort(values)
                    theoretical_quantiles = stats.norm.ppf(np.linspace(0.01, 0.99, len(sorted_values)))
                    
                    fig.add_trace(go.Scatter(
                        x=theoretical_quantiles,
                        y=sorted_values,
                        mode='markers',
                        name='Data',
                        marker=dict(color='#667eea', size=6)
                    ))
                    
                    # Reference line
                    fig.add_trace(go.Scatter(
                        x=[theoretical_quantiles.min(), theoretical_quantiles.max()],
                        y=[sorted_values.min(), sorted_values.max()],
                        mode='lines',
                        name='Normal Distribution',
                        line=dict(color='red', dash='dash')
                    ))
                    
                    fig.update_layout(
                        title=f"Q-Q Plot: {selected_num}",
                        xaxis_title="Theoretical Quantiles",
                        yaxis_title="Sample Quantiles"
                    )
                    st.plotly_chart(fig, use_container_width=True)
                    
                    st.markdown("""
                    **Cara Membaca Q-Q Plot:**
                    - Jika titik mengikuti garis merah → Data normal
                    - Jika titik melengkung ke atas → Right-skewed (ekor kanan panjang)
                    - Jika titik melengkung ke bawah → Left-skewed (ekor kiri panjang)
                    - Titik menyebar jauh dari garis → Tidak normal
                    """)
                
                # Tabel statistik lengkap untuk semua variabel numerik
                st.markdown("### Ringkasan Semua Variabel Numerik")
                
                desc_stats = df[num_cols].describe().T
                desc_stats['missing'] = df[num_cols].isnull().sum().values
                desc_stats['missing_pct'] = (desc_stats['missing'] / len(df) * 100).round(2)
                desc_stats = desc_stats.round(2)
                
                st.dataframe(desc_stats, use_container_width=True)
                
                # Download button untuk statistik deskriptif
                csv_desc = desc_stats.to_csv().encode('utf-8')
                st.download_button(
                    label="Download Statistik Deskriptif (CSV)",
                    data=csv_desc,
                    file_name=f"statistik_deskriptif_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
        
        # ==================== STATISTIK KATEGORIKAL ====================
        if len(cat_cols) > 0:
            st.markdown("---")
            st.subheader("Statistik Variabel Kategorikal")
            
            selected_cat = st.selectbox("Pilih variabel kategorikal untuk analisis detail:", cat_cols)
            
            if selected_cat:
                value_counts = df[selected_cat].value_counts()
                value_pct = df[selected_cat].value_counts(normalize=True) * 100
                
                col1, col2, col3 = st.columns(3)
                col1.metric("Unique Values", len(value_counts))
                col2.metric("Most Common", value_counts.index[0])
                col3.metric("Most Common %", f"{value_pct.iloc[0]:.1f}%")
                
                # Tabel frekuensi
                st.markdown("### Tabel Frekuensi")
                freq_df = pd.DataFrame({
                    'Kategori': value_counts.index,
                    'Jumlah': value_counts.values,
                    'Persentase (%)': value_pct.values.round(2)
                })
                st.dataframe(freq_df.head(20), use_container_width=True)
                
                if len(value_counts) > 20:
                    st.info(f"Menampilkan 20 dari {len(value_counts)} kategori")
                
                # Visualisasi
                tab1, tab2 = st.tabs(["Bar Chart", "Pie Chart"])
                
                with tab1:
                    # Ambil top 15 untuk visualisasi
                    top_n = min(15, len(value_counts))
                    fig = px.bar(
                        x=value_counts.index[:top_n],
                        y=value_counts.values[:top_n],
                        title=f"Top {top_n} Kategori: {selected_cat}",
                        labels={'x': selected_cat, 'y': 'Jumlah'},
                        text=value_counts.values[:top_n]
                    )
                    fig.update_traces(
                        marker_color='#667eea',
                        texttemplate='%{text}',
                        textposition='outside'
                    )
                    st.plotly_chart(fig, use_container_width=True)
                
                with tab2:
                    # Pie chart untuk top 10
                    top_n = min(10, len(value_counts))
                    fig = px.pie(
                        names=value_counts.index[:top_n],
                        values=value_counts.values[:top_n],
                        title=f"Proporsi Top {top_n}: {selected_cat}"
                    )
                    st.plotly_chart(fig, use_container_width=True)
        
        # ==================== STATISTIK TEMPORAL ====================
        if len(date_cols) > 0:
            st.markdown("---")
            st.subheader("Statistik Variabel Temporal")
            
            selected_date = st.selectbox("Pilih variabel tanggal untuk analisis:", date_cols)
            
            if selected_date:
                date_series = pd.to_datetime(df[selected_date].dropna())
                
                if len(date_series) > 0:
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("Tanggal Awal", date_series.min().strftime('%Y-%m-%d'))
                    col2.metric("Tanggal Akhir", date_series.max().strftime('%Y-%m-%d'))
                    col3.metric("Rentang (hari)", (date_series.max() - date_series.min()).days)
                    col4.metric("Total Data", len(date_series))
                    
                    # Time series trend
                    st.markdown("### Distribusi Data Sepanjang Waktu")
                    
                    # Agregasi per bulan
                    monthly_counts = date_series.dt.to_period('M').value_counts().sort_index()
                    monthly_counts.index = monthly_counts.index.to_timestamp()
                    
                    fig = go.Figure()
                    fig.add_trace(go.Scatter(
                        x=monthly_counts.index,
                        y=monthly_counts.values,
                        mode='lines+markers',
                        name='Jumlah Data',
                        line=dict(color='#667eea', width=2),
                        marker=dict(size=6)
                    ))
                    fig.update_layout(
                        title=f"Jumlah Data per Bulan: {selected_date}",
                        xaxis_title="Bulan",
                        yaxis_title="Jumlah Data",
                        hovermode='x unified'
                    )
                    st.plotly_chart(fig, use_container_width=True)
        
        # ==================== KORELASI ANTAR VARIABEL NUMERIK ====================
        if len(num_cols) >= 2:
            st.markdown("---")
            st.subheader("Matriks Korelasi (Exploratory)")
            
            st.markdown("""
            <div class="info-box">
                <strong>Correlation Matrix</strong><br>
                Melihat hubungan awal antar variabel numerik. Korelasi tinggi (mendekati ±1) menunjukkan hubungan kuat.
            </div>
            """, unsafe_allow_html=True)
            
            # Hitung korelasi
            corr_matrix = df[num_cols].corr()
            
            # Heatmap
            fig = px.imshow(
                corr_matrix,
                labels=dict(color="Korelasi"),
                x=num_cols,
                y=num_cols,
                color_continuous_scale='RdBu_r',
                aspect="auto",
                zmin=-1,
                zmax=1,
                text_auto='.2f'
            )
            fig.update_layout(
                title="Correlation Heatmap",
                height=600
            )
            st.plotly_chart(fig, use_container_width=True)
            
            # Tabel korelasi tertinggi
            st.markdown("### Top 10 Korelasi Terkuat")
            
            # Extract correlation pairs
            corr_pairs = []
            for i in range(len(num_cols)):
                for j in range(i + 1, len(num_cols)):
                    corr_pairs.append({
                        'Variabel 1': num_cols[i],
                        'Variabel 2': num_cols[j],
                        'Korelasi': corr_matrix.iloc[i, j]
                    })
            
            corr_df = pd.DataFrame(corr_pairs)
            corr_df['Abs Korelasi'] = corr_df['Korelasi'].abs()
            corr_df = corr_df.sort_values('Abs Korelasi', ascending=False).head(10)
            corr_df = corr_df[['Variabel 1', 'Variabel 2', 'Korelasi']].round(3)
            
            st.dataframe(corr_df, use_container_width=True)
        
        # ==================== NEXT STEPS ====================
        st.markdown("---")
        st.markdown("""
        <div class="success-box">
            <h3>Statistik Deskriptif Selesai!</h3>
            <p><strong>Langkah Selanjutnya:</strong></p>
            <ul>
                <li>Lanjut ke halaman <strong>"Analisis"</strong> untuk uji statistik inferensial</li>
                <li>Gunakan insight dari statistik deskriptif untuk memilih analisis yang tepat</li>
                <li>Perhatikan variabel dengan missing values atau outlier sebelum analisis lanjutan</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

# ==================== PAGE 3: ANALISIS (Enhanced) ====================
elif page == "Analisis":
    st.header("Analisis Data")
    
    if st.session_state.df is None:
        st.warning("Belum ada data. Upload dan terapkan tipe data di 'Input Data' dulu.")
    else:
        df = st.session_state.df
        types = st.session_state.column_types
        num_cols, date_cols, cat_cols = get_columns_by_type(df, types)
        
        st.markdown("""
        <div class="info-box">
            <strong>Pilih jenis analisis yang sesuai dengan pertanyaan Anda</strong>
        </div>
        """, unsafe_allow_html=True)
        
        analysis_type = st.selectbox(
            "Apa yang ingin Anda ketahui?",
            [
                "-- Pilih Tujuan Analisis --",
                "Apakah ada hubungan antara dua faktor? (Korelasi)",
                "Apakah ada perbedaan antar kelompok? (Perbandingan)",
                "Bagaimana pola perubahan dari waktu ke waktu? (Tren)",
                "Bisa memprediksi Y dari X? (Regresi Linear)",
                "Ada data yang tidak wajar? (Deteksi Outlier)",
                "ANALISIS LANJUTAN - Lihat di bawah"
            ]
        )
        
        # Advanced Analysis Section
        if "LANJUTAN" in analysis_type:
            st.markdown("---")
            st.markdown("### Analisis Lanjutan")
            st.markdown("""
            <div class="info-box">
                <strong>Pilih analisis statistik yang lebih advanced untuk insight mendalam</strong>
            </div>
            """, unsafe_allow_html=True)
            
            advanced_type = st.selectbox(
                "Pilih Jenis Analisis Lanjutan:",
                [
                    "-- Pilih Analisis Lanjutan --",
                    "Regresi Berganda (Prediksi dari Multiple X)",
                    "Chi-Square (Hubungan Kategori vs Kategori)",
                    "Forecasting (Prediksi Nilai Masa Depan)",
                    "Correlation Matrix (Korelasi Multi-Variabel)",
                    "Uji Normalitas Lengkap",
                    "Uji Homogenitas Varians (Levene's Test)"
                ]
            )
            
            # Override analysis_type dengan advanced_type
            if "Regresi Berganda" in advanced_type:
                analysis_type = advanced_type
            elif "Chi-Square" in advanced_type:
                analysis_type = advanced_type
            elif "Forecasting" in advanced_type:
                analysis_type = advanced_type
            elif "Correlation Matrix" in advanced_type:
                analysis_type = advanced_type
            elif "Normalitas" in advanced_type:
                analysis_type = advanced_type
            elif "Homogenitas" in advanced_type:
                analysis_type = advanced_type
        
        st.markdown("---")
        
        # KORELASI
        if "hubungan" in analysis_type.lower():
            st.subheader("Analisis Korelasi")
            
            st.markdown("""
            **Contoh pertanyaan:** Apakah semakin banyak patroli, semakin banyak temuan pelanggaran?
            """)
            
            if len(num_cols) < 2:
                st.error("Membutuhkan minimal 2 kolom angka untuk analisis ini")
            else:
                col1, col2 = st.columns(2)
                with col1:
                    x_var = st.selectbox("Variabel 1 (X)", num_cols, key="corr_x")
                with col2:
                    y_var = st.selectbox("Variabel 2 (Y)", [c for c in num_cols if c != x_var], key="corr_y")
                
                if st.button("Jalankan Analisis Korelasi", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis data..."):
                        result = analyze_correlation(df, x_var, y_var)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            analysis_data = {
                                'type': 'correlation',
                                'x_var': x_var,
                                'y_var': y_var,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.session_state.analysis_results = analysis_data
                            st.session_state.all_analysis_history.append(analysis_data)
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
        
        # PERBANDINGAN KELOMPOK
        elif "perbedaan" in analysis_type.lower():
            st.subheader("Analisis Perbandingan Kelompok")
            
            st.markdown("""
            **Contoh pertanyaan:** Apakah kinerja Stasiun A berbeda dengan Stasiun B?
            """)
            
            if len(cat_cols) == 0 or len(num_cols) == 0:
                st.error("Membutuhkan minimal 1 kolom kategori dan 1 kolom angka")
            else:
                col1, col2 = st.columns(2)
                with col1:
                    group_var = st.selectbox("Kelompok Pembanding", cat_cols, key="comp_group")
                with col2:
                    value_var = st.selectbox("Nilai yang Dibandingkan", num_cols, key="comp_value")
                
                if st.button("Jalankan Analisis Perbandingan", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis data..."):
                        result = analyze_group_comparison(df, group_var, value_var)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            analysis_data = {
                                'type': 'comparison',
                                'group_var': group_var,
                                'value_var': value_var,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.session_state.analysis_results = analysis_data
                            st.session_state.all_analysis_history.append(analysis_data)
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
        
        # TREN WAKTU
        elif "waktu" in analysis_type.lower():
            st.subheader("Analisis Tren Waktu")
            
            st.markdown("""
            **Contoh pertanyaan:** Apakah jumlah pelanggaran meningkat dari bulan ke bulan?
            """)
            
            if len(date_cols) == 0 or len(num_cols) == 0:
                st.error("Membutuhkan minimal 1 kolom tanggal dan 1 kolom angka")
            else:
                col1, col2, col3 = st.columns(3)
                with col1:
                    date_var = st.selectbox("Kolom Waktu", date_cols, key="trend_date")
                with col2:
                    value_var = st.selectbox("Nilai yang Dianalisis", num_cols, key="trend_value")
                with col3:
                    freq = st.selectbox("Ringkas Per", 
                                       ["Harian", "Mingguan", "Bulanan", "Triwulan", "Tahunan"],
                                       index=2)
                
                freq_map = {
                    "Harian": "D",
                    "Mingguan": "W",
                    "Bulanan": "M",
                    "Triwulan": "Q",
                    "Tahunan": "Y"
                }
                
                if st.button("Jalankan Analisis Tren", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis data..."):
                        result = analyze_trend(df, date_var, value_var, freq_map[freq])
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            analysis_data = {
                                'type': 'trend',
                                'date_var': date_var,
                                'value_var': value_var,
                                'freq': freq,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.session_state.analysis_results = analysis_data
                            st.session_state.all_analysis_history.append(analysis_data)
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
        
        # REGRESI LINEAR (NEW)
        elif "prediksi" in analysis_type.lower():
            st.subheader("Analisis Regresi Linear")
            
            st.markdown("""
            **Contoh pertanyaan:** Bisa prediksi jumlah pelanggaran berdasarkan jumlah patroli?
            
            Regresi linear mencari persamaan garis terbaik: Y = a + bX
            """)
            
            if len(num_cols) < 2:
                st.error("Membutuhkan minimal 2 kolom angka untuk analisis ini")
            else:
                col1, col2 = st.columns(2)
                with col1:
                    x_var = st.selectbox("Variabel Prediktor (X)", num_cols, key="reg_x")
                with col2:
                    y_var = st.selectbox("Variabel Target (Y)", [c for c in num_cols if c != x_var], key="reg_y")
                
                if st.button("Jalankan Analisis Regresi", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis data..."):
                        result = analyze_regression(df, x_var, y_var)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            st.session_state.analysis_results = {
                                'type': 'regression',
                                'x_var': x_var,
                                'y_var': y_var,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
                            
                            # Quick preview
                            st.markdown("### Preview Hasil")
                            col1, col2, col3, col4 = st.columns(4)
                            col1.metric("R²", f"{result['r_squared']:.3f}")
                            col2.metric("Slope (b)", f"{result['slope']:.4f}")
                            col3.metric("Intercept (a)", f"{result['intercept']:.4f}")
                            col4.metric("P-value", f"{result['p_value']:.4f}")
                            
                            st.markdown(f"""
                            **Persamaan:** Y = {result['intercept']:.4f} + {result['slope']:.4f} × X
                            
                            **Interpretasi:** Jika X naik 1 unit, Y akan {'naik' if result['slope'] > 0 else 'turun'} 
                            sebesar {abs(result['slope']):.4f} unit.
                            """)
        
        # DETEKSI OUTLIER (NEW)
        elif "wajar" in analysis_type.lower():
            st.subheader("Deteksi Outlier")
            
            st.markdown("""
            **Tujuan:** Menemukan data yang tidak biasa atau ekstrem
            
            **Metode yang digunakan:**
            - IQR (Interquartile Range): Deteksi nilai di luar Q1-1.5×IQR hingga Q3+1.5×IQR
            - Z-score: Deteksi nilai dengan Z-score > 3 (sangat jauh dari rata-rata)
            """)
            
            if len(num_cols) == 0:
                st.error("Membutuhkan minimal 1 kolom angka untuk analisis ini")
            else:
                var_col = st.selectbox("Pilih Variabel untuk Deteksi Outlier", num_cols, key="outlier_var")
                
                if st.button("Deteksi Outlier", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis data..."):
                        result = analyze_outliers(df, var_col)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            st.session_state.analysis_results = {
                                'type': 'outlier',
                                'var_col': var_col,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
                            
                            # Quick preview
                            st.markdown("### Preview Hasil")
                            col1, col2, col3 = st.columns(3)
                            col1.metric("Total Data", f"{result['n_total']:,}")
                            col2.metric("Outlier (IQR)", f"{result['n_outliers_iqr']:,} ({result['pct_outliers_iqr']:.1f}%)")
                            col3.metric("Outlier (Z-score)", f"{result['n_outliers_zscore']:,} ({result['pct_outliers_zscore']:.1f}%)")
                            
                            if result['n_outliers_iqr'] > 0:
                                st.warning(f"Ditemukan {result['n_outliers_iqr']} outlier (metode IQR)")
                            else:
                                st.success("Tidak ada outlier terdeteksi (metode IQR)")
        
        # ==================== ADVANCED ANALYSIS SECTIONS ====================
        
        # REGRESI BERGANDA
        elif "Regresi Berganda" in analysis_type:
            st.subheader("Analisis Regresi Berganda")
            
            st.markdown("""
            **Tujuan:** Prediksi variabel Y menggunakan beberapa variabel X sekaligus
            
            **Contoh:** Prediksi jumlah pelanggaran dari: jumlah patroli, cuaca, dan musim
            
            **Persamaan:** Y = a + b₁X₁ + b₂X₂ + b₃X₃ + ...
            """)
            
            if len(num_cols) < 3:
                st.error("Membutuhkan minimal 3 kolom angka (2 prediktor + 1 target)")
            else:
                st.markdown("**Pilih Variabel:**")
                
                y_var = st.selectbox("Variabel Target (Y) - yang ingin diprediksi:", num_cols, key="mreg_y")
                
                available_x = [c for c in num_cols if c != y_var]
                x_vars = st.multiselect(
                    "Variabel Prediktor (X) - pilih 2 atau lebih:",
                    available_x,
                    default=available_x[:min(2, len(available_x))],
                    key="mreg_x"
                )
                
                if len(x_vars) < 2:
                    st.warning("Pilih minimal 2 variabel prediktor untuk regresi berganda")
                elif st.button("Jalankan Regresi Berganda", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis data..."):
                        result = analyze_multiple_regression(df, x_vars, y_var)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            st.session_state.analysis_results = {
                                'type': 'multiple_regression',
                                'x_vars': x_vars,
                                'y_var': y_var,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
                            
                            # Quick preview
                            st.markdown("### Preview Hasil")
                            col1, col2, col3 = st.columns(3)
                            col1.metric("R²", f"{result['r2']:.3f}")
                            col2.metric("Adjusted R²", f"{result['adj_r2']:.3f}")
                            col3.metric("RMSE", f"{result['rmse']:.4f}")
                            
                            st.markdown("**Koefisien:**")
                            coef_df = pd.DataFrame({
                                'Variabel': list(result['coefficients'].keys()),
                                'Koefisien': list(result['coefficients'].values())
                            })
                            st.dataframe(coef_df, use_container_width=True)
        
        # CHI-SQUARE TEST
        elif "Chi-Square" in analysis_type:
            st.subheader("Uji Chi-Square")
            
            st.markdown("""
            **Tujuan:** Menguji hubungan antara dua variabel kategorikal
            
            **Contoh pertanyaan:**
            - Apakah jenis pelanggaran berkaitan dengan wilayah?
            - Apakah metode patroli berpengaruh pada tingkat keberhasilan?
            
            **Metode:** Chi-Square Test of Independence + Cramér's V untuk ukuran efek
            """)
            
            if len(cat_cols) < 2:
                st.error("Membutuhkan minimal 2 kolom kategorikal")
            else:
                col1, col2 = st.columns(2)
                with col1:
                    var1 = st.selectbox("Variabel Kategori 1:", cat_cols, key="chi_var1")
                with col2:
                    var2 = st.selectbox("Variabel Kategori 2:", [c for c in cat_cols if c != var1], key="chi_var2")
                
                if st.button("Jalankan Uji Chi-Square", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis data..."):
                        result = analyze_chi_square(df, var1, var2)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            st.session_state.analysis_results = {
                                'type': 'chi_square',
                                'var1': var1,
                                'var2': var2,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
                            
                            # Quick preview
                            st.markdown("### Preview Hasil")
                            col1, col2, col3, col4 = st.columns(4)
                            col1.metric("Chi-Square", f"{result['chi2']:.2f}")
                            col2.metric("P-value", f"{result['p_value']:.4f}")
                            col3.metric("Cramér's V", f"{result['cramers_v']:.3f}")
                            col4.metric("Kekuatan", result['strength'])
                            
                            st.markdown("**Tabel Kontingensi:**")
                            st.dataframe(result['contingency_table'], use_container_width=True)
        
        # TIME SERIES FORECASTING
        elif "Forecasting" in analysis_type:
            st.subheader("Forecasting (Prediksi Masa Depan)")
            
            st.markdown("""
            **Tujuan:** Memprediksi nilai di masa depan berdasarkan pola historis
            
            **Metode:** Linear trend forecasting
            
            **Cocok untuk:** Data dengan pola tren yang jelas
            """)
            
            if len(date_cols) == 0 or len(num_cols) == 0:
                st.error("Membutuhkan minimal 1 kolom tanggal dan 1 kolom angka")
            else:
                col1, col2, col3 = st.columns(3)
                with col1:
                    date_var = st.selectbox("Kolom Tanggal:", date_cols, key="fc_date")
                with col2:
                    value_var = st.selectbox("Variabel untuk Forecast:", num_cols, key="fc_value")
                with col3:
                    n_periods = st.number_input("Berapa periode ke depan?", min_value=1, max_value=12, value=3)
                
                if st.button("Jalankan Forecasting", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis dan memprediksi..."):
                        result = analyze_time_series_forecast(df, date_var, value_var, n_periods)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            st.session_state.analysis_results = {
                                'type': 'forecast',
                                'date_var': date_var,
                                'value_var': value_var,
                                'n_periods': n_periods,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
                            
                            # Quick preview
                            st.markdown("### Preview Hasil")
                            col1, col2, col3 = st.columns(3)
                            col1.metric("R² (Model Fit)", f"{result['r_squared']:.3f}")
                            col2.metric("RMSE", f"{result['rmse']:.4f}")
                            col3.metric("Periode Prediksi", n_periods)
                            
                            st.markdown("**Prediksi:**")
                            forecast_df = pd.DataFrame({
                                'Tanggal': [d.strftime('%Y-%m-%d') for d in result['future_dates']],
                                'Nilai Prediksi': result['forecast_values']
                            })
                            st.dataframe(forecast_df, use_container_width=True)
        
        # CORRELATION MATRIX
        elif "Correlation Matrix" in analysis_type:
            st.subheader("Correlation Matrix (Multi-Variabel)")
            
            st.markdown("""
            **Tujuan:** Lihat korelasi antara semua variabel numerik sekaligus
            
            **Kegunaan:** 
            - Identifikasi hubungan tersembunyi
            - Deteksi multikolinearitas
            - Eksplorasi data awal
            """)
            
            if len(num_cols) < 2:
                st.error("Membutuhkan minimal 2 kolom numerik")
            else:
                selected_vars = st.multiselect(
                    "Pilih variabel untuk analisis (pilih 2 atau lebih):",
                    num_cols,
                    default=num_cols[:min(5, len(num_cols))],
                    key="corr_matrix_vars"
                )
                
                if len(selected_vars) < 2:
                    st.warning("Pilih minimal 2 variabel")
                elif st.button("Jalankan Analisis Correlation Matrix", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis korelasi..."):
                        result = analyze_correlation_matrix(df, selected_vars)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            st.session_state.analysis_results = {
                                'type': 'correlation_matrix',
                                'variables': selected_vars,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
                            
                            # Quick preview - heatmap
                            st.markdown("### Preview: Correlation Heatmap")
                            fig = px.imshow(
                                result['corr_matrix'],
                                labels=dict(color="Korelasi"),
                                x=selected_vars,
                                y=selected_vars,
                                color_continuous_scale='RdBu_r',
                                aspect="auto",
                                zmin=-1,
                                zmax=1
                            )
                            fig.update_layout(title="Correlation Matrix Heatmap")
                            st.plotly_chart(fig, use_container_width=True)
                            
                            # Top correlations
                            st.markdown("**Top 5 Korelasi Terkuat:**")
                            st.dataframe(result['corr_pairs'].head(5), use_container_width=True)
        
        # NORMALITY TEST
        elif "Normalitas" in analysis_type:
            st.subheader("Uji Normalitas Lengkap")
            
            st.markdown("""
            **Tujuan:** Menguji apakah data mengikuti distribusi normal
            
            **Pentingnya:** Banyak uji statistik parametrik mengasumsikan data normal
            
            **Metode yang digunakan:**
            - Shapiro-Wilk Test
            - Kolmogorov-Smirnov Test
            - Anderson-Darling Test
            - Skewness & Kurtosis
            """)
            
            if len(num_cols) == 0:
                st.error("Membutuhkan minimal 1 kolom numerik")
            else:
                var_col = st.selectbox("Pilih Variabel:", num_cols, key="norm_var")
                
                if st.button("Jalankan Uji Normalitas", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis distribusi..."):
                        result = analyze_normality_test(df, var_col)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            st.session_state.analysis_results = {
                                'type': 'normality',
                                'var_col': var_col,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
                            
                            # Quick preview
                            st.markdown("### Preview Hasil")
                            col1, col2, col3 = st.columns(3)
                            col1.metric("Shapiro p-value", f"{result['shapiro_p']:.4f}")
                            col2.metric("K-S p-value", f"{result['ks_p']:.4f}")
                            col3.metric("Skewness", f"{result['skewness']:.3f}")
                            
                            if result['is_normal_sw'] and result['is_normal_ks']:
                                st.success("Data mengikuti distribusi normal")
                            else:
                                st.warning("Data tidak mengikuti distribusi normal")
        
        # VARIANCE HOMOGENEITY TEST
        elif "Homogenitas" in analysis_type:
            st.subheader("Uji Homogenitas Varians")
            
            st.markdown("""
            **Tujuan:** Menguji apakah varians antar kelompok sama (homogen)
            
            **Pentingnya:** Asumsi penting untuk t-test dan ANOVA
            
            **Metode:** Levene's Test (robust terhadap non-normalitas)
            """)
            
            if len(cat_cols) == 0 or len(num_cols) == 0:
                st.error("Membutuhkan minimal 1 kolom kategori dan 1 kolom angka")
            else:
                col1, col2 = st.columns(2)
                with col1:
                    group_var = st.selectbox("Variabel Kelompok:", cat_cols, key="lev_group")
                with col2:
                    value_var = st.selectbox("Variabel Nilai:", num_cols, key="lev_value")
                
                if st.button("Jalankan Uji Homogenitas", type="primary", use_container_width=True):
                    with st.spinner("Menganalisis homogenitas varians..."):
                        result = analyze_variance_homogeneity(df, group_var, value_var)
                        
                        if 'error' in result:
                            st.error(f"{result['error']}")
                        else:
                            st.session_state.analysis_results = {
                                'type': 'homogeneity',
                                'group_var': group_var,
                                'value_var': value_var,
                                'result': result,
                                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
                            st.success("Analisis selesai! Lihat tab 'Hasil & Export'")
                            
                            # Quick preview
                            st.markdown("### Preview Hasil")
                            col1, col2, col3 = st.columns(3)
                            col1.metric("Levene Statistic", f"{result['stat']:.4f}")
                            col2.metric("P-value", f"{result['p_value']:.4f}")
                            col3.metric("Jumlah Kelompok", result['n_groups'])
                            
                            if result['is_homogeneous']:
                                st.success(f"{result['conclusion']}")
                            else:
                                st.warning(f"{result['conclusion']}")
                            
                            st.info(f"**Rekomendasi:** {result['recommendation']}")

# ==================== PAGE 4: HASIL & EXPORT (Enhanced) ====================
elif page == "Hasil & Export":
    st.header("Hasil Analisis & Export")
    
    if st.session_state.analysis_results is None:
        st.info("Belum ada hasil. Jalankan analisis di tab 'Analisis' dulu.")
    else:
        res = st.session_state.analysis_results
        
        # Display results based on type
        if res['type'] == 'correlation':
            r = res['result']
            st.subheader(f"Hubungan: {res['x_var']} vs {res['y_var']}")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Korelasi (r)", f"{r['r']:.3f}")
            col2.metric("P-value", f"{r['p_value']:.4f}")
            col3.metric("Jumlah Data", f"{r['n']:,}")
            col4.metric("Metode", r['method'])
            
            fig = px.scatter(
                x=r['x'], y=r['y'],
                labels={'x': res['x_var'], 'y': res['y_var']},
                title=f"Scatter Plot: {res['x_var']} vs {res['y_var']}",
                trendline="ols",
                trendline_color_override="red"
            )
            fig.update_traces(marker=dict(size=8, opacity=0.6, color='#667eea'))
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### Interpretasi")
            st.markdown(f"""
            Terdapat hubungan {r['direction']} dengan kekuatan {r['strength']} (r = {r['r']:.3f}).
            
            Hubungan ini {r['significance']} secara statistik (p = {r['p_value']:.4f}, alpha = 0.05).
            
            **Arti Praktis:** {r['practical']}
            """)
            
            with st.expander("Uji Asumsi & Metode Statistik"):
                st.markdown(f"""
                <div class="assumption-box">
                    <strong>Metode yang Digunakan: {r['method']} Correlation</strong><br>
                    {r['method_explain']}<br><br>
                    
                    <strong>Uji Normalitas (Shapiro-Wilk):</strong><br>
                    • {res['x_var']}: p = {r['p_norm_x']:.4f} {'Normal' if r['p_norm_x'] > 0.05 else 'Tidak Normal'}<br>
                    • {res['y_var']}: p = {r['p_norm_y']:.4f} {'Normal' if r['p_norm_y'] > 0.05 else 'Tidak Normal'}<br><br>
                    
                    <strong>Kenapa {r['method']}?</strong><br>
                    {'Data kedua variabel mendekati distribusi normal, sehingga Pearson (linear correlation) adalah pilihan yang tepat.' if r['is_normal'] else 'Data tidak normal, sehingga Spearman (rank-based correlation) lebih robust dan cocok untuk data ini.'}
                </div>
                """, unsafe_allow_html=True)
        
        elif res['type'] == 'comparison':
            r = res['result']
            st.subheader(f"Perbandingan: {res['value_var']} per {res['group_var']}")
            
            col1, col2, col3 = st.columns(3)
            col1.metric("Uji Statistik", f"{r['stat']:.3f}")
            col2.metric("P-value", f"{r['p_value']:.4f}")
            col3.metric("Jumlah Kelompok", r['n_groups'])
            
            fig = px.bar(
                r['summary'], 
                x=res['group_var'], 
                y='Rata-rata',
                title=f"Rata-rata {res['value_var']} per {res['group_var']}",
                text='Rata-rata',
                color='Rata-rata',
                color_continuous_scale='Blues'
            )
            fig.update_traces(texttemplate='%{text:.2f}', textposition='outside')
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### Statistik per Kelompok")
            st.dataframe(r['summary'], use_container_width=True)
            
            st.markdown("### Interpretasi")
            sorted_df = r['summary'].sort_values('Rata-rata', ascending=False)
            top = sorted_df.iloc[0]
            bottom = sorted_df.iloc[-1]
            
            st.markdown(f"""
            Kelompok "{top[res['group_var']]}" memiliki rata-rata tertinggi ({top['Rata-rata']:.2f}), 
            sedangkan "{bottom[res['group_var']]}" memiliki rata-rata terendah ({bottom['Rata-rata']:.2f}).
            
            Perbedaan antar kelompok {r['significance']} secara statistik (p = {r['p_value']:.4f}).
            
            {'Artinya: Perbedaan yang terlihat bukan hanya kebetulan, ada indikasi kuat bahwa kelompok-kelompok ini memang berbeda.' if r['p_value'] < 0.05 else 'Artinya: Perbedaan yang terlihat bisa jadi karena variasi acak, belum cukup bukti untuk menyimpulkan ada perbedaan yang konsisten.'}
            """)
        
        elif res['type'] == 'trend':
            r = res['result']
            st.subheader(f"Tren: {res['value_var']} dari Waktu ke Waktu")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Nilai Awal", f"{r['first_val']:.2f}")
            col2.metric("Nilai Akhir", f"{r['last_val']:.2f}")
            col3.metric("Perubahan", f"{r['change']:+.2f} ({r['change_pct']:+.1f}%)")
            col4.metric("Periode", r['n_periods'])
            
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=r['dates'], 
                y=r['values'],
                mode='lines+markers',
                name=res['value_var'],
                line=dict(color='#667eea', width=3),
                marker=dict(size=8)
            ))
            fig.update_layout(
                title=f"Tren {res['value_var']} ({res['freq']})",
                xaxis_title="Periode",
                yaxis_title=res['value_var'],
                hovermode='x unified'
            )
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### Interpretasi")
            st.markdown(f"""
            Tren menunjukkan pola {r['trend_desc']} dengan kekuatan {r['trend_strength']} (r = {r['r']:.3f}, p = {r['p_value']:.4f}).
            
            Dari periode awal ke akhir, terjadi perubahan sebesar {r['change']:+.2f} ({r['change_pct']:+.1f}%).
            
            {'Artinya: Ada pola perubahan yang konsisten dari waktu ke waktu, bukan hanya fluktuasi acak.' if r['p_value'] < 0.05 else 'Artinya: Polanya belum cukup konsisten untuk disebut tren yang jelas, mungkin lebih ke fluktuasi.'}
            """)
        
        elif res['type'] == 'regression':
            r = res['result']
            st.subheader(f"Regresi Linear: Prediksi {res['y_var']} dari {res['x_var']}")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("R² (Goodness of Fit)", f"{r['r_squared']:.3f}")
            col2.metric("Slope (b)", f"{r['slope']:.4f}")
            col3.metric("Intercept (a)", f"{r['intercept']:.4f}")
            col4.metric("RMSE", f"{r['rmse']:.4f}")
            
            # Scatter plot dengan garis regresi
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=r['x'], y=r['y'],
                mode='markers',
                name='Data Aktual',
                marker=dict(size=8, opacity=0.6, color='#667eea')
            ))
            fig.add_trace(go.Scatter(
                x=r['x'], y=r['y_pred'],
                mode='lines',
                name='Garis Regresi',
                line=dict(color='red', width=3)
            ))
            fig.update_layout(
                title=f"Regresi: {res['y_var']} = {r['intercept']:.4f} + {r['slope']:.4f} × {res['x_var']}",
                xaxis_title=res['x_var'],
                yaxis_title=res['y_var'],
                hovermode='closest'
            )
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### Interpretasi")
            st.markdown(f"""
            **Persamaan Regresi:**  
            {res['y_var']} = {r['intercept']:.4f} + {r['slope']:.4f} × {res['x_var']}
            
            **Arti Koefisien:**
            - Slope ({r['slope']:.4f}): Jika {res['x_var']} naik 1 unit, {res['y_var']} akan {'naik' if r['slope'] > 0 else 'turun'} sebesar {abs(r['slope']):.4f} unit
            - Intercept ({r['intercept']:.4f}): Nilai prediksi {res['y_var']} ketika {res['x_var']} = 0
            
            **Akurasi Model:**
            - R² = {r['r_squared']:.3f}: Model menjelaskan {r['r_squared']*100:.1f}% variasi dalam data
            - RMSE = {r['rmse']:.4f}: Error rata-rata prediksi
            - Model ini {'signifikan' if r['p_value'] < 0.05 else 'tidak signifikan'} (p = {r['p_value']:.4f})
            
            **Kesimpulan:**  
            {'Model cukup baik untuk prediksi (R² > 0.5)' if r['r_squared'] > 0.5 else 'Model lemah untuk prediksi (R² < 0.5), pertimbangkan faktor lain'}
            """)
            
            # Residual plot
            st.markdown("### Analisis Residual")
            fig_resid = go.Figure()
            fig_resid.add_trace(go.Scatter(
                x=r['y_pred'], y=r['residuals'],
                mode='markers',
                marker=dict(size=8, opacity=0.6, color='#764ba2')
            ))
            fig_resid.add_hline(y=0, line_dash="dash", line_color="red")
            fig_resid.update_layout(
                title="Residual Plot (untuk cek asumsi)",
                xaxis_title="Nilai Prediksi",
                yaxis_title="Residual (Error)",
                hovermode='closest'
            )
            st.plotly_chart(fig_resid, use_container_width=True)
            
            st.info("""
            **Cara Baca Residual Plot:**
            - Jika titik tersebar acak di sekitar garis 0 → Model baik
            - Jika ada pola (kurva, corong) → Model kurang cocok
            """)
        
        elif res['type'] == 'outlier':
            r = res['result']
            st.subheader(f"Deteksi Outlier: {res['var_col']}")
            
            col1, col2, col3 = st.columns(3)
            col1.metric("Total Data", f"{r['n_total']:,}")
            col2.metric("Outlier (IQR)", f"{r['n_outliers_iqr']:,} ({r['pct_outliers_iqr']:.1f}%)")
            col3.metric("Outlier (Z-score)", f"{r['n_outliers_zscore']:,} ({r['pct_outliers_zscore']:.1f}%)")
            
            # Box plot
            fig = go.Figure()
            fig.add_trace(go.Box(
                y=r['all_values'],
                name=res['var_col'],
                marker_color='#667eea',
                boxmean='sd'
            ))
            fig.update_layout(
                title=f"Box Plot: {res['var_col']} (Deteksi Outlier)",
                yaxis_title=res['var_col'],
                showlegend=False
            )
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### Interpretasi")
            
            if r['n_outliers_iqr'] > 0:
                st.warning(f"""
                **Outlier Terdeteksi (Metode IQR)**
                
                - Jumlah: {r['n_outliers_iqr']:,} data ({r['pct_outliers_iqr']:.1f}% dari total)
                - Batas bawah: {r['lower_bound']:.2f}
                - Batas atas: {r['upper_bound']:.2f}
                
                **Nilai Outlier:**
                {', '.join([f'{x:.2f}' for x in sorted(r['outliers_iqr'])[:10]])}
                {f'...dan {len(r["outliers_iqr"])-10} lainnya' if len(r['outliers_iqr']) > 10 else ''}
                """)
            else:
                st.success("Tidak ada outlier terdeteksi (metode IQR)")
            
            if r['n_outliers_zscore'] > 0:
                st.info(f"""
                **Outlier Ekstrem (Z-score > 3)**
                
                - Jumlah: {r['n_outliers_zscore']:,} data ({r['pct_outliers_zscore']:.1f}% dari total)
                - Nilai sangat jauh dari rata-rata (>3 standar deviasi)
                
                **Nilai Outlier Ekstrem:**
                {', '.join([f'{x:.2f}' for x in sorted(r['outliers_zscore'])[:10]])}
                """)
            
            st.markdown("""
            **Tindak Lanjut:**
            1. Periksa apakah outlier adalah kesalahan input data
            2. Jika valid, investigasi kenapa nilai ini ekstrem
            3. Pertimbangkan untuk mengecualikan outlier dari analisis tertentu
            4. Atau gunakan metode statistik yang robust terhadap outlier
            """)
        
        # MULTIPLE REGRESSION
        elif res['type'] == 'multiple_regression':
            r = res['result']
            st.subheader(f"Regresi Berganda: Prediksi {res['y_var']}")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("R²", f"{r['r2']:.3f}")
            col2.metric("Adjusted R²", f"{r['adj_r2']:.3f}")
            col3.metric("RMSE", f"{r['rmse']:.4f}")
            col4.metric("MAE", f"{r['mae']:.4f}")
            
            # Equation
            equation_parts = [f"{r['intercept']:.4f}"]
            for var, coef in r['coefficients'].items():
                equation_parts.append(f"{coef:+.4f} × {var}")
            equation = f"{res['y_var']} = " + " ".join(equation_parts)
            
            st.markdown(f"""
            ### Persamaan Regresi
```
            {equation}
```
            """)
            
            # Coefficients table
            st.markdown("### Koefisien Regresi")
            coef_df = pd.DataFrame({
                'Variabel': ['Intercept'] + list(r['coefficients'].keys()),
                'Koefisien': [r['intercept']] + list(r['coefficients'].values()),
                'Interpretasi': ['Nilai dasar'] + [
                    f"Jika naik 1 unit, Y {'naik' if c > 0 else 'turun'} {abs(c):.4f}"
                    for c in r['coefficients'].values()
                ]
            })
            st.dataframe(coef_df, use_container_width=True)
            
            # Scatter: actual vs predicted
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=r['y_actual'], 
                y=r['y_pred'],
                mode='markers',
                name='Prediksi',
                marker=dict(size=8, opacity=0.6, color='#667eea')
            ))
            fig.add_trace(go.Scatter(
                x=[r['y_actual'].min(), r['y_actual'].max()],
                y=[r['y_actual'].min(), r['y_actual'].max()],
                mode='lines',
                name='Perfect Fit',
                line=dict(color='red', dash='dash')
            ))
            fig.update_layout(
                title="Actual vs Predicted Values",
                xaxis_title="Nilai Aktual",
                yaxis_title="Nilai Prediksi",
                hovermode='closest'
            )
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### Interpretasi")
            st.markdown(f"""
            **Akurasi Model:**
            - R² = {r['r2']:.3f}: Model menjelaskan {r['r2']*100:.1f}% variasi dalam {res['y_var']}
            - Adjusted R² = {r['adj_r2']:.3f}: Memperhitungkan jumlah prediktor
            - RMSE = {r['rmse']:.4f}: Error rata-rata prediksi
            
            **Variabel Prediktor:** {', '.join(res['x_vars'])}
            
            **Kesimpulan:**
            {'Model sangat baik untuk prediksi (R² > 0.7)' if r['r2'] > 0.7 else 'Model cukup baik (0.5 < R² ≤ 0.7)' if r['r2'] > 0.5 else 'Model lemah (R² ≤ 0.5), pertimbangkan variabel lain'}
            """)
        
        # CHI-SQUARE
        elif res['type'] == 'chi_square':
            r = res['result']
            st.subheader(f"Chi-Square: {res['var1']} vs {res['var2']}")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Chi-Square", f"{r['chi2']:.2f}")
            col2.metric("P-value", f"{r['p_value']:.4f}")
            col3.metric("Cramér's V", f"{r['cramers_v']:.3f}")
            col4.metric("Degrees of Freedom", r['dof'])
            
            # Contingency table
            st.markdown("### Tabel Kontingensi (Crosstab)")
            st.dataframe(r['contingency_table'], use_container_width=True)
            
            # Heatmap
            fig = px.imshow(
                r['contingency_table'],
                labels=dict(x=res['var2'], y=res['var1'], color="Frekuensi"),
                x=r['contingency_table'].columns,
                y=r['contingency_table'].index,
                color_continuous_scale='Blues',
                aspect="auto"
            )
            fig.update_layout(title=f"Heatmap: {res['var1']} vs {res['var2']}")
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### Interpretasi")
            st.markdown(f"""
            **Hasil Uji:**
            - Chi-Square statistic = {r['chi2']:.2f}
            - P-value = {r['p_value']:.4f}
            - Hubungan {r['significance']} secara statistik
            
            **Ukuran Efek (Cramér's V):**
            - Nilai = {r['cramers_v']:.3f}
            - Kekuatan = {r['strength']}
            
            **Kesimpulan:**
            {'Ada hubungan yang signifikan antara kedua variabel kategorikal ini.' if r['p_value'] < 0.05 else 'Tidak ada cukup bukti hubungan antara kedua variabel.'}
            """)
        
        # FORECASTING
        elif res['type'] == 'forecast':
            r = res['result']
            st.subheader(f"Forecasting: {res['value_var']}")
            
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("R² (Model Fit)", f"{r['r_squared']:.3f}")
            col2.metric("RMSE", f"{r['rmse']:.4f}")
            col3.metric("Data Historis", r['n_historical'])
            col4.metric("Periode Forecast", res['n_periods'])
            
            # Time series plot dengan forecast
            fig = go.Figure()
            
            # Historical data
            fig.add_trace(go.Scatter(
                x=r['historical_dates'],
                y=r['historical_values'],
                mode='lines+markers',
                name='Data Aktual',
                line=dict(color='#667eea', width=2),
                marker=dict(size=6)
            ))
            
            # Fitted values
            fig.add_trace(go.Scatter(
                x=r['historical_dates'],
                y=r['fitted_values'],
                mode='lines',
                name='Model Fit',
                line=dict(color='orange', width=2, dash='dash')
            ))
            
            # Forecast
            fig.add_trace(go.Scatter(
                x=r['future_dates'],
                y=r['forecast_values'],
                mode='lines+markers',
                name='Forecast',
                line=dict(color='red', width=2),
                marker=dict(size=8, symbol='star')
            ))
            
            fig.update_layout(
                title=f"Forecasting: {res['value_var']}",
                xaxis_title="Tanggal",
                yaxis_title=res['value_var'],
                hovermode='x unified',
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            st.plotly_chart(fig, use_container_width=True)
            
            # Forecast table
            st.markdown("### Prediksi Nilai Masa Depan")
            forecast_df = pd.DataFrame({
                'Tanggal': [d.strftime('%Y-%m-%d') for d in r['future_dates']],
                'Nilai Prediksi': [f"{v:.2f}" for v in r['forecast_values']]
            })
            st.dataframe(forecast_df, use_container_width=True)
            
            st.markdown("### Interpretasi")
            st.markdown(f"""
            **Kualitas Model:**
            - R² = {r['r_squared']:.3f}: Model menjelaskan {r['r_squared']*100:.1f}% variasi data historis
            - RMSE = {r['rmse']:.4f}: Error rata-rata dari data historis
            
            **Trend:**
            - Slope = {r['slope']:.4f}
            - {'Tren naik' if r['slope'] > 0 else 'Tren turun'} sebesar {abs(r['slope']):.4f} per periode
            
            **Catatan Penting:**
            - Forecasting ini menggunakan linear trend sederhana
            - Akurasi bergantung pada pola data historis
            - Semakin jauh prediksi, semakin rendah akurasinya
            """)
        
        # CORRELATION MATRIX
        elif res['type'] == 'correlation_matrix':
            r = res['result']
            st.subheader("Correlation Matrix")
            
            col1, col2 = st.columns(2)
            col1.metric("Jumlah Variabel", len(res['variables']))
            col2.metric("Total Data", f"{r['n']:,}")
            
            # Heatmap
            st.markdown("### Correlation Heatmap")
            fig = px.imshow(
                r['corr_matrix'],
                labels=dict(color="Korelasi"),
                x=res['variables'],
                y=res['variables'],
                color_continuous_scale='RdBu_r',
                aspect="auto",
                zmin=-1,
                zmax=1,
                text_auto='.2f'
            )
            fig.update_layout(
                title="Correlation Matrix",
                height=600
            )
            st.plotly_chart(fig, use_container_width=True)
            
            # Top correlations
            st.markdown("### Korelasi Terkuat")
            top_corr = r['corr_pairs'].head(10)
            st.dataframe(top_corr, use_container_width=True)
            
            st.markdown("### Interpretasi")
            
            # Find strongest positive and negative
            strongest = r['corr_pairs'].iloc[0]
            st.markdown(f"""
            **Korelasi Terkuat:**
            - {strongest['var1']} vs {strongest['var2']}
            - Korelasi = {strongest['correlation']:.3f}
            """)
        
        # NORMALITY TEST
        elif res['type'] == 'normality':
            r = res['result']
            st.subheader(f"Uji Normalitas: {res['var_col']}")
            
            col1, col2, col3 = st.columns(3)
            col1.metric("Shapiro p-value", f"{r['shapiro_p']:.4f}")
            col2.metric("K-S p-value", f"{r['ks_p']:.4f}")
            col3.metric("Jumlah Data", f"{r['n']:,}")
            
            # Histogram with normal curve
            st.markdown("### Histogram vs Kurva Normal")
            
            fig = go.Figure()
            
            # Histogram
            fig.add_trace(go.Histogram(
                x=r['values'],
                nbinsx=30,
                name='Data Aktual',
                marker_color='#667eea',
                opacity=0.7,
                histnorm='probability density'
            ))
            
            # Normal curve
            mean = r['values'].mean()
            std = r['values'].std()
            x_range = np.linspace(r['values'].min(), r['values'].max(), 100)
            normal_curve = stats.norm.pdf(x_range, mean, std)
            
            fig.add_trace(go.Scatter(
                x=x_range,
                y=normal_curve,
                mode='lines',
                name='Kurva Normal Teoritis',
                line=dict(color='red', width=2)
            ))
            
            fig.update_layout(
                title=f"Distribusi {res['var_col']}",
                xaxis_title=res['var_col'],
                yaxis_title="Density",
                showlegend=True
            )
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### Interpretasi")
            
            if r['conclusion'] == 'Data mengikuti distribusi normal':
                st.success(f"{r['conclusion']}")
            else:
                st.warning(f"{r['conclusion']}")
        
        # HOMOGENEITY TEST  
        elif res['type'] == 'homogeneity':
            r = res['result']
            st.subheader(f"Uji Homogenitas Varians")
            
            col1, col2, col3 = st.columns(3)
            col1.metric("Levene Statistic", f"{result['stat']:.4f}")
            col2.metric("P-value", f"{r['p_value']:.4f}")
            col3.metric("Jumlah Kelompok", r['n_groups'])
            
            st.markdown("### Interpretasi")
            
            if r['is_homogeneous']:
                st.success(f"{r['conclusion']}")
            else:
                st.warning(f"{r['conclusion']}")
            
            st.info(f"**Rekomendasi:** {r['recommendation']}")
        
        # ==================== EXPORT SECTION ====================
        st.markdown("---")
        st.markdown("""
        <div class="download-section">
            <h3>Download Hasil Analisis</h3>
            <p>Pilih format yang Anda inginkan untuk menyimpan hasil analisis</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### Export ke CSV")
            st.markdown("""
            Download hasil analisis dalam format CSV (mudah dibuka di Excel, Google Sheets, dll)
            """)
            
            csv_summary, csv_detail, csv_stats = export_to_csv(res, res['type'])
            
            if csv_summary is not None:
                csv1 = csv_summary.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Download Ringkasan (CSV)",
                    data=csv1,
                    file_name=f"ringkasan_{res['type']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            if csv_detail is not None:
                csv2 = csv_detail.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Download Data Detail (CSV)",
                    data=csv2,
                    file_name=f"data_detail_{res['type']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            if csv_stats is not None:
                csv3 = csv_stats.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Download Statistik (CSV)",
                    data=csv3,
                    file_name=f"statistik_{res['type']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
        
        with col2:
            st.markdown("### Export ke Excel")
            st.markdown("""
            Download hasil dalam format Excel
            
            **Keuntungan:**
            - File Excel dengan multiple sheets
            - Format profesional dengan warna
            - Bisa diedit dan disesuaikan
            """)
            
            if st.button("Generate Excel File", use_container_width=True):
                with st.spinner("Membuat file Excel..."):
                    excel_file = create_excel_with_formulas(res, res['type'])
                    
                    st.download_button(
                        label="Download Excel (dengan Formula)",
                        data=excel_file,
                        file_name=f"hasil_analisis_{res['type']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    st.success("File Excel siap didownload!")
                    
                    st.info("""
                    **Cara Menggunakan File Excel:**
                    1. Buka file di Excel atau Google Sheets
                    2. Lihat multiple sheets untuk informasi lengkap
                    3. Gunakan untuk laporan atau presentasi
                    """)

# ==================== PAGE 5: KESIMPULAN ====================
elif page == "Kesimpulan":
    st.header("Kesimpulan & Rekomendasi")
    
    if st.session_state.analysis_results is None:
        st.info("Belum ada hasil untuk disimpulkan. Jalankan analisis dulu.")
    else:
        res = st.session_state.analysis_results
        
        st.markdown("""
        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    padding: 1.5rem; border-radius: 10px; color: white; margin-bottom: 2rem;'>
            <h2 style='color: white; margin: 0;'>Ringkasan Temuan Utama</h2>
        </div>
        """, unsafe_allow_html=True)
        
        if res['type'] == 'correlation':
            r = res['result']
            
            st.markdown(f"""
            **Jenis Analisis:** Hubungan Antar Variabel (Korelasi {r['method']})
            
            **Variabel:** `{res['x_var']}` vs `{res['y_var']}`
            
            **Temuan Kunci:**
            - Korelasi: {r['r']:.3f} ({r['strength']}, {r['direction']})
            - Signifikansi: {r['significance']} (p = {r['p_value']:.4f})
            - Berdasarkan {r['n']:,} pasang data
            - Waktu analisis: {res.get('timestamp', 'N/A')}
            """)
            
            st.markdown("### Rekomendasi Tindak Lanjut")
            
            if abs(r['r']) > 0.5 and r['p_value'] < 0.05:
                st.markdown("""
                <div class="success-box">
                    <strong>Hubungan Kuat Terdeteksi</strong><br><br>
                    <strong>Yang Bisa Dilakukan:</strong><br>
                    1. Strategi: Fokus peningkatan pada variabel X untuk mempengaruhi variabel Y<br>
                    2. Prediksi: Gunakan variabel X sebagai indikator untuk memperkirakan Y<br>
                    3. Monitoring: Pantau kedua variabel secara bersamaan untuk deteksi dini<br>
                    4. Analisis Lanjut: Cari faktor penyebab kenapa kedua variabel berhubungan<br>
                    5. Validasi: Uji hubungan ini di periode atau wilayah lain untuk konfirmasi<br>
                    6. Gunakan Regresi Linear untuk membuat model prediksi
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("""
                <div class="info-box">
                    <strong>Hubungan Lemah/Tidak Signifikan</strong><br><br>
                    <strong>Yang Perlu Dipertimbangkan:</strong><br>
                    1. Cari variabel lain yang mungkin lebih berpengaruh<br>
                    2. Periksa apakah ada variabel moderator<br>
                    3. Evaluasi kualitas data - apakah ada noise atau kesalahan pengukuran?<br>
                    4. Pertimbangkan hubungan non-linear<br>
                    5. Cek outlier yang mungkin mempengaruhi hasil
                </div>
                """, unsafe_allow_html=True)
        
        elif res['type'] == 'comparison':
            r = res['result']
            sorted_df = r['summary'].sort_values('Rata-rata', ascending=False)
            top = sorted_df.iloc[0]
            bottom = sorted_df.iloc[-1]
            
            st.markdown(f"""
            **Jenis Analisis:** Perbandingan Kelompok ({r['test_name']})
            
            **Variabel:** `{res['value_var']}` dibandingkan antar `{res['group_var']}`
            
            **Temuan Kunci:**
            - Kelompok terbaik: {top[res['group_var']]} (rata-rata {top['Rata-rata']:.2f})
            - Kelompok terendah: {bottom[res['group_var']]} (rata-rata {bottom['Rata-rata']:.2f})
            - Perbedaan: {r['significance']} (p = {r['p_value']:.4f})
            - Total kelompok dibandingkan: {r['n_groups']}
            - Waktu analisis: {res.get('timestamp', 'N/A')}
            """)
            
            st.markdown("### Rekomendasi Tindak Lanjut")
            
            if r['p_value'] < 0.05:
                st.markdown(f"""
                <div class="success-box">
                    <strong>Ada Perbedaan Signifikan Antar Kelompok</strong><br><br>
                    <strong>Langkah Strategis:</strong><br>
                    1. Investigasi: Cari tahu kenapa {top[res['group_var']]} unggul - apa best practice yang bisa ditiru?<br>
                    2. Benchmarking: Jadikan {top[res['group_var']]} sebagai benchmark untuk kelompok lain<br>
                    3. Alokasi Sumber Daya: Evaluasi apakah kelompok dengan performa rendah butuh dukungan tambahan<br>
                    4. Analisis Akar Masalah: Untuk {bottom[res['group_var']]}, cari tahu apa kendala utamanya<br>
                    5. Target Improvement: Set target realistis untuk kelompok menengah-bawah untuk naik level<br>
                    6. Dokumentasi: Catat best practice dari top performer untuk replikasi
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("""
                <div class="info-box">
                    <strong>Perbedaan Tidak Signifikan</strong><br><br>
                    <strong>Interpretasi:</strong><br>
                    Meski ada perbedaan angka, secara statistik perbedaan ini bisa jadi karena variasi acak.<br><br>
                    <strong>Yang Bisa Dilakukan:</strong><br>
                    1. Jika perbedaan terlihat kecil, ini bisa jadi hal baik (konsistensi antar kelompok)<br>
                    2. Evaluasi apakah standar/SOP sudah diterapkan merata<br>
                    3. Pertimbangkan faktor lain yang mungkin lebih membedakan kelompok<br>
                    4. Tambah ukuran sample jika memungkinkan
                </div>
                """, unsafe_allow_html=True)
        
        elif res['type'] == 'trend':
            r = res['result']
            
            st.markdown(f"""
            **Jenis Analisis:** Tren Waktu
            
            **Variabel:** `{res['value_var']}` dari waktu ke waktu
            
            **Temuan Kunci:**
            - Tren: {r['trend_desc']} ({r['trend_strength']})
            - Perubahan: {r['change']:+.2f} ({r['change_pct']:+.1f}%)
            - Nilai awal: {r['first_val']:.2f}, Nilai akhir: {r['last_val']:.2f}
            - Signifikansi: p = {r['p_value']:.4f}
            - Periode: {res['freq']}
            - Waktu analisis: {res.get('timestamp', 'N/A')}
            """)
            
            st.markdown("### Rekomendasi Tindak Lanjut")
            
            if r['p_value'] < 0.05:
                if r['trend_desc'] == 'naik':
                    st.markdown("""
                    <div class="success-box">
                        <strong>Tren Naik Terdeteksi</strong><br><br>
                        <strong>Langkah Strategis:</strong><br>
                        1. Identifikasi faktor pendorong tren positif ini<br>
                        2. Perkuat program/kegiatan yang mendukung tren naik<br>
                        3. Monitor apakah tren akan berkelanjutan atau mencapai plateau<br>
                        4. Dokumentasikan best practice untuk replikasi<br>
                        5. Set target lebih tinggi jika tren stabil<br>
                        6. Waspadai potential ceiling effect
                    </div>
                    """, unsafe_allow_html=True)
                elif r['trend_desc'] == 'turun':
                    st.markdown("""
                    <div class="warning-box">
                        <strong>Tren Turun Terdeteksi</strong><br><br>
                        <strong>Langkah Strategis:</strong><br>
                        1. Investigasi penyebab penurunan - apakah faktor internal atau eksternal?<br>
                        2. Evaluasi program/kebijakan yang mungkin perlu diperbaiki<br>
                        3. Set target untuk membalikkan tren negatif<br>
                        4. Monitor progress secara berkala<br>
                        5. Benchmark dengan periode sebelumnya atau unit lain<br>
                        6. Eskalasi jika tren terus memburuk
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.markdown("""
                <div class="info-box">
                    <strong>Tren Tidak Jelas/Stabil</strong><br><br>
                    <strong>Interpretasi:</strong><br>
                    Data menunjukkan fluktuasi tanpa pola naik/turun yang konsisten.<br><br>
                    <strong>Yang Bisa Dilakukan:</strong><br>
                    1. Stabilitas bisa jadi hal positif - menunjukkan konsistensi<br>
                    2. Periksa apakah ada pola musiman atau siklikal<br>
                    3. Evaluasi apakah perlu intervensi untuk meningkatkan performa<br>
                    4. Pertimbangkan analisis dengan periode yang lebih panjang
                </div>
                """, unsafe_allow_html=True)
        
        # ==================== HISTORY SECTION ====================
        if len(st.session_state.all_analysis_history) > 1:
            st.markdown("---")
            st.markdown("### Riwayat Analisis")
            
            history_df = pd.DataFrame([
                {
                    'Waktu': h.get('timestamp', 'N/A'),
                    'Jenis': h['type'].title(),
                    'Variabel': f"{h.get('x_var', h.get('group_var', h.get('date_var', h.get('var_col', 'N/A'))))}"
                }
                for h in st.session_state.all_analysis_history
            ])
            
            st.dataframe(history_df, use_container_width=True)

# ==================== FOOTER ====================
st.sidebar.markdown("---")
st.sidebar.markdown(f"""
<div style='font-size: 0.8rem; color: #666;'>
    <strong>Dashboard Analisis PSDKP</strong><br>
</div>
""", unsafe_allow_html=True)
